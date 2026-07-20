"""
OSV Demand Scheduler Local Server
=================================
A Flask-based server that supports:
- Serving static files (HTML, CSS, JS)
- API endpoints for reading/writing data
- Excel workbook upload and parsing

Run with: python server.py
Access at: http://127.0.0.1:8000
"""

import os
import json
import uuid
import re
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename

# Try to import openpyxl for Excel parsing
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("WARNING: openpyxl not installed. Excel upload will be limited.")
    print("Install with: pip install openpyxl")

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# Configuration
DATA_DIR = Path(__file__).parent / 'data'
SNAPSHOTS_DIR = DATA_DIR / 'snapshots'
UPLOAD_FOLDER = Path(__file__).parent / 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'json'}

# Ensure directories exist
DATA_DIR.mkdir(exist_ok=True)
SNAPSHOTS_DIR.mkdir(exist_ok=True)
UPLOAD_FOLDER.mkdir(exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def generate_id():
    return uuid.uuid4().hex[:12]


def parse_date(value):
    """Parse various date formats to ISO string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common formats
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S']:
            try:
                return datetime.strptime(value, fmt).isoformat()
            except ValueError:
                continue
    return str(value) if value else None


def safe_string(value):
    """Convert value to string safely."""
    if value is None:
        return ''
    return str(value).strip()


def safe_float(value, default=0.0):
    """Convert value to float safely."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


# ============== Static File Routes ==============

@app.route('/')
def index():
    return send_file('index.html')


@app.route('/spot-hire')
@app.route('/spot_hire.html')
def spot_hire():
    return send_file('spot_hire.html')


@app.route('/<path:filename>')
def serve_static(filename):
    # Serve static files from root directory
    if os.path.isfile(filename):
        return send_file(filename)
    return send_from_directory('.', filename)


# ============== Health Check ==============

@app.route('/api/health')
def health():
    return jsonify({
        'status': 'ok',
        'excel_support': EXCEL_SUPPORT,
        'timestamp': datetime.now().isoformat()
    })


# ============== Data API Routes ==============

@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    try:
        with open(DATA_DIR / 'tasks.json', 'r') as f:
            return jsonify(json.load(f))
    except FileNotFoundError:
        return jsonify({'tasks': [], 'source': '', 'buffer_hours': 24})


@app.route('/api/tasks', methods=['POST'])
def save_tasks():
    data = request.get_json()
    with open(DATA_DIR / 'tasks.json', 'w') as f:
        json.dump(data, f, indent=2)
    return jsonify({'success': True})


@app.route('/api/conflicts', methods=['GET'])
def get_conflicts():
    try:
        with open(DATA_DIR / 'conflicts.json', 'r') as f:
            return jsonify(json.load(f))
    except FileNotFoundError:
        return jsonify({'conflicts': [], 'fleet': None})


@app.route('/api/conflicts', methods=['POST'])
def save_conflicts():
    data = request.get_json()
    with open(DATA_DIR / 'conflicts.json', 'w') as f:
        json.dump(data, f, indent=2)
    return jsonify({'success': True})


@app.route('/api/asset-capacity', methods=['GET'])
def get_asset_capacity():
    try:
        with open(DATA_DIR / 'asset-capacity.json', 'r') as f:
            return jsonify(json.load(f))
    except FileNotFoundError:
        return jsonify({'asset_capacities': [], 'monthly_capacity_entries': []})


@app.route('/api/asset-capacity', methods=['POST'])
def save_asset_capacity():
    data = request.get_json()
    with open(DATA_DIR / 'asset-capacity.json', 'w') as f:
        json.dump(data, f, indent=2)
    return jsonify({'success': True})


@app.route('/api/spot-hire', methods=['GET'])
def get_spot_hire():
    try:
        with open(DATA_DIR / 'spot-hire.json', 'r') as f:
            return jsonify(json.load(f))
    except FileNotFoundError:
        return jsonify({'records': [], 'source': '', 'phase_colors': {}})


@app.route('/api/spot-hire', methods=['POST'])
def save_spot_hire():
    data = request.get_json()
    
    # Check if this is a single new record (has 'asset' and 'activity' but no 'records')
    if 'asset' in data and 'activity' in data and 'records' not in data:
        # This is a new record to add - load existing data and append
        try:
            with open(DATA_DIR / 'spot-hire.json', 'r') as f:
                spot_data = json.load(f)
        except FileNotFoundError:
            spot_data = {'records': [], 'source': '', 'phase_colors': {}}
        
        # Generate ID for new record
        new_record = {
            'id': generate_id(),
            'asset': data.get('asset', ''),
            'display_asset': data.get('display_asset') or data.get('asset', ''),
            'vessel_count': data.get('vessel_count', 1),
            'area': data.get('area', ''),
            'activity': data.get('activity', ''),
            'phase': data.get('phase', ''),
            'color': data.get('color', '#78909c'),
            'start_date': data.get('start_date', ''),
            'end_date': data.get('end_date', ''),
            'status': data.get('status', 'Planned'),
            'notes': data.get('notes', ''),
            'source': 'app'
        }
        spot_data.setdefault('records', []).append(new_record)
        
        with open(DATA_DIR / 'spot-hire.json', 'w') as f:
            json.dump(spot_data, f, indent=2)
        update_asset_capacity_from_spot_hire(spot_data.get('records', []))
        return jsonify({'success': True, 'id': new_record['id']})
    
    # Otherwise, save the full data structure
    with open(DATA_DIR / 'spot-hire.json', 'w') as f:
        json.dump(data, f, indent=2)
    update_asset_capacity_from_spot_hire(data.get('records', []))
    return jsonify({'success': True})


# Spot hire impacts/notes per month
IMPACTS_FILE = DATA_DIR / 'spot-hire-impacts.json'

def load_impacts():
    try:
        with open(IMPACTS_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_impacts_file(data):
    with open(IMPACTS_FILE, 'w') as f:
        json.dump(data, f, indent=2)


@app.route('/api/spot-hire/impacts', methods=['GET'])
def get_spot_hire_impacts():
    month = request.args.get('month', '')
    impacts = load_impacts()
    if month and month in impacts:
        return jsonify(impacts[month])
    return jsonify({
        'base_fleet': '',
        'frac_spot_hires': '',
        'operational_spot_hires': '',
        'text': '',
        'updated_at': None
    })


@app.route('/api/spot-hire/impacts', methods=['POST'])
def save_spot_hire_impacts():
    data = request.get_json()
    month = data.get('month', '')
    if not month:
        return jsonify({'error': 'Month required'}), 400
    
    impacts = load_impacts()
    impacts[month] = {
        'base_fleet': data.get('base_fleet', ''),
        'frac_spot_hires': data.get('frac_spot_hires', ''),
        'operational_spot_hires': data.get('operational_spot_hires', ''),
        'text': data.get('text', ''),
        'updated_at': datetime.now().isoformat()
    }
    save_impacts_file(impacts)
    return jsonify(impacts[month])


@app.route('/api/spot-hire/changes', methods=['POST'])
def apply_spot_hire_changes():
    """Apply batch changes to spot hire records (update dates, etc)."""
    data = request.get_json()
    changes = data.get('changes', [])
    
    # Load current data
    try:
        with open(DATA_DIR / 'spot-hire.json', 'r') as f:
            spot_data = json.load(f)
    except FileNotFoundError:
        spot_data = {'records': [], 'source': '', 'phase_colors': {}}
    
    records = spot_data.get('records', [])
    records_by_id = {r['id']: r for r in records}
    
    # Apply changes
    for change in changes:
        record_id = change.get('id')
        if record_id and record_id in records_by_id:
            record = records_by_id[record_id]
            # Update all editable fields
            for field in ['start_date', 'end_date', 'status', 'phase', 'notes', 
                          'vessel_count', 'asset', 'display_asset', 'area', 
                          'activity', 'color']:
                if field in change:
                    record[field] = change[field]
    
    # Save
    spot_data['records'] = list(records_by_id.values())
    with open(DATA_DIR / 'spot-hire.json', 'w') as f:
        json.dump(spot_data, f, indent=2)
    update_asset_capacity_from_spot_hire(spot_data.get('records', []))
    
    return jsonify({'success': True, 'changes_applied': len(changes)})


@app.route('/api/spot-hire/<record_id>', methods=['DELETE'])
def delete_spot_hire_record(record_id):
    """Delete a spot hire record by ID."""
    try:
        with open(DATA_DIR / 'spot-hire.json', 'r') as f:
            spot_data = json.load(f)
    except FileNotFoundError:
        return jsonify({'error': 'Record not found'}), 404
    
    records = spot_data.get('records', [])
    original_count = len(records)
    spot_data['records'] = [r for r in records if r.get('id') != record_id]
    
    if len(spot_data['records']) == original_count:
        return jsonify({'error': 'Record not found'}), 404
    
    with open(DATA_DIR / 'spot-hire.json', 'w') as f:
        json.dump(spot_data, f, indent=2)
    update_asset_capacity_from_spot_hire(spot_data.get('records', []))
    
    return jsonify({'success': True})


# ============== Snapshot History API ==============

@app.route('/api/snapshots', methods=['GET'])
def list_snapshots():
    """List all available task snapshots for planned vs actual tracking."""
    snapshots = []
    for snapshot_file in sorted(SNAPSHOTS_DIR.glob('tasks_*.json'), reverse=True):
        try:
            with open(snapshot_file, 'r') as f:
                data = json.load(f)
            snapshots.append({
                'id': snapshot_file.name,
                'date': data.get('snapshot_date', ''),
                'time': data.get('snapshot_time', ''),
                'imported_at': data.get('imported_at', ''),
                'source': data.get('source', ''),
                'task_count': len(data.get('tasks', []))
            })
        except (json.JSONDecodeError, IOError):
            continue
    return jsonify({'snapshots': snapshots})


@app.route('/api/snapshots/<snapshot_id>', methods=['GET'])
def get_snapshot(snapshot_id):
    """Get a specific snapshot by ID."""
    snapshot_path = SNAPSHOTS_DIR / snapshot_id
    if not snapshot_path.exists():
        return jsonify({'error': 'Snapshot not found'}), 404
    try:
        with open(snapshot_path, 'r') as f:
            return jsonify(json.load(f))
    except (json.JSONDecodeError, IOError) as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/snapshots/compare', methods=['GET'])
def compare_snapshots():
    """Compare two snapshots to identify new, removed, and changed tasks.
    
    Query params:
    - baseline: snapshot_id of the baseline (older) snapshot
    - current: snapshot_id of the current (newer) snapshot, or 'current' for live data
    - asset: optional asset filter
    
    Returns tasks grouped by: new (late additions), removed, changed, unchanged
    """
    baseline_id = request.args.get('baseline')
    current_id = request.args.get('current', 'current')
    asset_filter = request.args.get('asset')
    
    if not baseline_id:
        return jsonify({'error': 'baseline parameter required'}), 400
    
    # Load baseline snapshot
    baseline_path = SNAPSHOTS_DIR / baseline_id
    if not baseline_path.exists():
        return jsonify({'error': f'Baseline snapshot {baseline_id} not found'}), 404
    
    try:
        with open(baseline_path, 'r') as f:
            baseline_data = json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        return jsonify({'error': f'Error reading baseline: {e}'}), 500
    
    # Load current data (either from snapshot or live)
    if current_id == 'current':
        try:
            with open(DATA_DIR / 'tasks.json', 'r') as f:
                current_data = json.load(f)
        except FileNotFoundError:
            current_data = {'tasks': []}
    else:
        current_path = SNAPSHOTS_DIR / current_id
        if not current_path.exists():
            return jsonify({'error': f'Current snapshot {current_id} not found'}), 404
        try:
            with open(current_path, 'r') as f:
                current_data = json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            return jsonify({'error': f'Error reading current: {e}'}), 500
    
    baseline_tasks = baseline_data.get('tasks', [])
    current_tasks = current_data.get('tasks', [])
    
    # Apply asset filter if specified
    if asset_filter:
        baseline_tasks = [t for t in baseline_tasks if t.get('asset') == asset_filter]
        current_tasks = [t for t in current_tasks if t.get('asset') == asset_filter]
    
    # Helper to parse dates for comparison
    def parse_task_date(date_str):
        """Parse a date string to a datetime object for comparison."""
        if not date_str:
            return None
        try:
            if 'T' in date_str:
                return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return datetime.strptime(date_str[:10], '%Y-%m-%d')
        except (ValueError, TypeError):
            return None
    
    # Create a soft key (asset + activity only) for grouping similar tasks
    def soft_key(t):
        """Generate a key based on asset and activity only (no date)."""
        asset = (t.get('asset') or '').strip().lower()
        activity = (t.get('activity') or t.get('project') or '').strip().lower()
        return f"{asset}|{activity}"
    
    # Group tasks by soft key
    from collections import defaultdict
    baseline_by_soft_key = defaultdict(list)
    for t in baseline_tasks:
        baseline_by_soft_key[soft_key(t)].append(t)
    
    current_by_soft_key = defaultdict(list)
    for t in current_tasks:
        current_by_soft_key[soft_key(t)].append(t)
    
    # Categorize tasks
    new_tasks = []  # In current but not baseline (late additions)
    removed_tasks = []  # In baseline but not current
    changed_tasks = []  # In both but with differences
    unchanged_tasks = []  # In both, identical
    matched_baseline_indices = set()  # Track matched baseline tasks by (soft_key, index)
    matched_current_indices = set()   # Track matched current tasks by (soft_key, index)
    
    # Fields to compare for changes (now includes start_date)
    compare_fields = ['start_date', 'status', 'offshore_start', 'offshore_end', 'return_end',
                      'duration_hours', 'transit_hours', 'vessel', 'coordinator']
    
    # Define change categories based on which fields changed
    def categorize_changes(changes_dict):
        """Determine change categories based on which fields were modified."""
        categories = []
        date_fields = {'start_date', 'offshore_start', 'offshore_end', 'return_end'}
        duration_fields = {'duration_hours', 'transit_hours'}
        
        changed_field_names = set(changes_dict.keys())
        
        if changed_field_names & date_fields:
            categories.append({'id': 'date_shift', 'label': 'Date Shift', 'icon': '📅', 'color': '#3182ce'})
        if changed_field_names & duration_fields:
            categories.append({'id': 'duration_change', 'label': 'Duration Change', 'icon': '⏱️', 'color': '#805ad5'})
        if 'status' in changed_field_names:
            categories.append({'id': 'status_update', 'label': 'Status Update', 'icon': '🔄', 'color': '#38a169'})
        if 'vessel' in changed_field_names:
            categories.append({'id': 'vessel_change', 'label': 'Vessel Change', 'icon': '🚢', 'color': '#dd6b20'})
        if 'coordinator' in changed_field_names:
            categories.append({'id': 'coordinator_change', 'label': 'Coordinator Change', 'icon': '👤', 'color': '#718096'})
        
        return categories
    
    # Track category counts for summary
    category_counts = {
        'date_shift': 0,
        'duration_change': 0,
        'status_update': 0,
        'vessel_change': 0,
        'coordinator_change': 0
    }
    
    # Match tasks using soft key (asset + activity), then pair by closest dates
    all_soft_keys = set(baseline_by_soft_key.keys()) | set(current_by_soft_key.keys())
    
    for sk in all_soft_keys:
        baseline_group = baseline_by_soft_key.get(sk, [])
        current_group = current_by_soft_key.get(sk, [])
        
        # Create pairing based on closest dates
        baseline_available = list(range(len(baseline_group)))
        current_available = list(range(len(current_group)))
        
        # Match each current task to the closest available baseline task
        for ci in list(current_available):
            current_task = current_group[ci]
            current_date = parse_task_date(current_task.get('start_date'))
            
            best_match_bi = None
            best_match_diff = float('inf')
            
            for bi in baseline_available:
                baseline_task = baseline_group[bi]
                baseline_date = parse_task_date(baseline_task.get('start_date'))
                
                if current_date and baseline_date:
                    diff = abs((current_date - baseline_date).days)
                else:
                    diff = 0 if (not current_date and not baseline_date) else 9999
                
                if diff < best_match_diff:
                    best_match_diff = diff
                    best_match_bi = bi
            
            if best_match_bi is not None and best_match_diff <= 30:  # Within 30 days = likely same route
                baseline_task = baseline_group[best_match_bi]
                baseline_available.remove(best_match_bi)
                current_available.remove(ci)
                matched_current_indices.add((sk, ci))
                matched_baseline_indices.add((sk, best_match_bi))
                
                # Compare fields
                changes = {}
                for field in compare_fields:
                    old_val = baseline_task.get(field)
                    new_val = current_task.get(field)
                    
                    # Normalize dates for comparison (just compare date portion)
                    if field in ['start_date', 'offshore_start', 'offshore_end', 'return_end']:
                        old_date = parse_task_date(old_val)
                        new_date = parse_task_date(new_val)
                        if old_date and new_date:
                            # Consider same if within 1 hour (handles timezone issues)
                            if abs((new_date - old_date).total_seconds()) < 3600:
                                continue
                    
                    if old_val != new_val:
                        if not old_val and not new_val:
                            continue
                        changes[field] = {'old': old_val, 'new': new_val}
                
                if changes:
                    categories = categorize_changes(changes)
                    for cat in categories:
                        if cat['id'] in category_counts:
                            category_counts[cat['id']] += 1
                    changed_tasks.append({
                        **current_task,
                        'change_type': 'changed',
                        'changes': changes,
                        'change_categories': categories,
                        'baseline_task': baseline_task
                    })
                else:
                    unchanged_tasks.append({
                        **current_task,
                        'change_type': 'unchanged'
                    })
        
        # Remaining current tasks are new
        for ci in current_available:
            new_tasks.append({
                **current_group[ci],
                'change_type': 'new',
                'first_seen': current_data.get('imported_at', '')
            })
        
        # Remaining baseline tasks are removed
        for bi in baseline_available:
            removed_tasks.append({
                **baseline_group[bi],
                'change_type': 'removed'
            })
    
    # Get unique assets for summary
    assets = sorted(set(t.get('asset', '') for t in current_tasks + baseline_tasks))
    
    return jsonify({
        'baseline': {
            'id': baseline_id,
            'date': baseline_data.get('snapshot_date', ''),
            'task_count': len(baseline_tasks)
        },
        'current': {
            'id': current_id,
            'date': current_data.get('snapshot_date', current_data.get('imported_at', '')[:10] if current_data.get('imported_at') else ''),
            'task_count': len(current_tasks)
        },
        'summary': {
            'new_count': len(new_tasks),
            'removed_count': len(removed_tasks),
            'changed_count': len(changed_tasks),
            'unchanged_count': len(unchanged_tasks),
            'assets': assets,
            'category_counts': category_counts
        },
        'new_tasks': new_tasks,
        'removed_tasks': removed_tasks,
        'changed_tasks': changed_tasks,
        'unchanged_tasks': unchanged_tasks
    })


@app.route('/api/snapshots/volatility', methods=['GET'])
def analyze_volatility():
    """Analyze schedule volatility/stability over time.
    
    Compares consecutive snapshots to measure how much the schedule changes.
    
    Query params:
        days: Number of days to analyze (default: 30)
        asset: Filter by specific asset (optional)
    """
    days = request.args.get('days', 30, type=int)
    asset_filter = request.args.get('asset', '').strip()
    
    # Get all snapshots sorted chronologically
    snapshot_files = sorted(SNAPSHOTS_DIR.glob('tasks_*.json'))
    
    if len(snapshot_files) < 2:
        return jsonify({
            'error': 'Need at least 2 snapshots for volatility analysis',
            'snapshot_count': len(snapshot_files)
        }), 400
    
    # Filter to requested time window
    cutoff_date = datetime.now() - timedelta(days=days)
    filtered_snapshots = []
    for sf in snapshot_files:
        try:
            # Parse date from filename: tasks_YYYYMMDD_HHMMSS.json
            date_str = sf.stem.split('_')[1]
            snap_date = datetime.strptime(date_str, '%Y%m%d')
            if snap_date >= cutoff_date:
                filtered_snapshots.append(sf)
        except (IndexError, ValueError):
            continue
    
    if len(filtered_snapshots) < 2:
        # Include at least 2 most recent snapshots
        filtered_snapshots = snapshot_files[-2:]
    
    # Compare consecutive snapshots
    comparisons = []
    total_new = 0
    total_removed = 0
    total_changed = 0
    total_unchanged = 0
    total_date_shifts = 0
    total_shift_days = 0
    shift_count = 0
    
    asset_volatility = {}  # Track volatility by asset
    weekly_data = {}  # Track by week
    
    def parse_date(d):
        if not d:
            return None
        try:
            if 'T' in str(d):
                return datetime.fromisoformat(str(d).replace('Z', '+00:00'))
            return datetime.strptime(str(d).split()[0], '%Y-%m-%d')
        except:
            return None
    
    for i in range(len(filtered_snapshots) - 1):
        baseline_path = filtered_snapshots[i]
        current_path = filtered_snapshots[i + 1]
        
        with open(baseline_path, 'r') as f:
            baseline_data = json.load(f)
        with open(current_path, 'r') as f:
            current_data = json.load(f)
        
        baseline_tasks = baseline_data.get('tasks', [])
        current_tasks = current_data.get('tasks', [])
        
        # Apply asset filter
        if asset_filter:
            baseline_tasks = [t for t in baseline_tasks if asset_filter.lower() in t.get('asset', '').lower()]
            current_tasks = [t for t in current_tasks if asset_filter.lower() in t.get('asset', '').lower()]
        
        # Simple matching by asset + activity (soft key)
        baseline_keys = {(t.get('asset', ''), t.get('activity', '')): t for t in baseline_tasks}
        current_keys = {(t.get('asset', ''), t.get('activity', '')): t for t in current_tasks}
        
        new_count = 0
        removed_count = 0
        changed_count = 0
        unchanged_count = 0
        period_shifts = 0
        
        # Find new and changed tasks
        for key, task in current_keys.items():
            asset = task.get('asset', 'Unassigned')
            if asset not in asset_volatility:
                asset_volatility[asset] = {'new': 0, 'removed': 0, 'changed': 0, 'unchanged': 0, 'total_shift_days': 0, 'shift_count': 0}
            
            if key in baseline_keys:
                baseline_task = baseline_keys[key]
                # Check for date changes
                old_start = parse_date(baseline_task.get('start_date'))
                new_start = parse_date(task.get('start_date'))
                
                if old_start and new_start and old_start != new_start:
                    changed_count += 1
                    asset_volatility[asset]['changed'] += 1
                    diff_days = abs((new_start - old_start).days)
                    if diff_days >= 1:
                        period_shifts += 1
                        total_shift_days += diff_days
                        shift_count += 1
                        asset_volatility[asset]['total_shift_days'] += diff_days
                        asset_volatility[asset]['shift_count'] += 1
                else:
                    unchanged_count += 1
                    asset_volatility[asset]['unchanged'] += 1
            else:
                new_count += 1
                asset_volatility[asset]['new'] += 1
        
        # Find removed tasks
        for key, task in baseline_keys.items():
            asset = task.get('asset', 'Unassigned')
            if asset not in asset_volatility:
                asset_volatility[asset] = {'new': 0, 'removed': 0, 'changed': 0, 'unchanged': 0, 'total_shift_days': 0, 'shift_count': 0}
            if key not in current_keys:
                removed_count += 1
                asset_volatility[asset]['removed'] += 1
        
        total_new += new_count
        total_removed += removed_count
        total_changed += changed_count
        total_unchanged += unchanged_count
        total_date_shifts += period_shifts
        
        # Track weekly data
        try:
            date_str = current_path.stem.split('_')[1]
            snap_date = datetime.strptime(date_str, '%Y%m%d')
            week_key = snap_date.strftime('%Y-W%W')
            if week_key not in weekly_data:
                weekly_data[week_key] = {'new': 0, 'removed': 0, 'changed': 0, 'total_tasks': 0, 'comparisons': 0}
            weekly_data[week_key]['new'] += new_count
            weekly_data[week_key]['removed'] += removed_count
            weekly_data[week_key]['changed'] += changed_count
            weekly_data[week_key]['total_tasks'] = len(current_tasks)
            weekly_data[week_key]['comparisons'] += 1
        except:
            pass
        
        comparisons.append({
            'baseline': baseline_path.name,
            'current': current_path.name,
            'new': new_count,
            'removed': removed_count,
            'changed': changed_count,
            'unchanged': unchanged_count,
            'date_shifts': period_shifts
        })
    
    # Calculate summary metrics
    total_comparisons = len(comparisons)
    total_tasks_tracked = total_new + total_removed + total_changed + total_unchanged
    
    # Volatility rate = (changes / total tasks) * 100
    volatility_rate = round((total_new + total_removed + total_changed) / max(total_tasks_tracked, 1) * 100, 1)
    
    # Average days shifted
    avg_shift_days = round(total_shift_days / max(shift_count, 1), 1)
    
    # Cancellation rate
    cancellation_rate = round(total_removed / max(total_tasks_tracked, 1) * 100, 1)
    
    # New route rate
    new_route_rate = round(total_new / max(total_tasks_tracked, 1) * 100, 1)
    
    # Process asset volatility into ranked list
    asset_rankings = []
    for asset, stats in asset_volatility.items():
        total_events = stats['new'] + stats['removed'] + stats['changed']
        total_asset_tasks = total_events + stats['unchanged']
        asset_vol_rate = round(total_events / max(total_asset_tasks, 1) * 100, 1)
        avg_asset_shift = round(stats['total_shift_days'] / max(stats['shift_count'], 1), 1) if stats['shift_count'] > 0 else 0
        asset_rankings.append({
            'asset': asset,
            'volatility_rate': asset_vol_rate,
            'new': stats['new'],
            'removed': stats['removed'],
            'changed': stats['changed'],
            'avg_shift_days': avg_asset_shift,
            'total_events': total_events
        })
    
    asset_rankings.sort(key=lambda x: x['volatility_rate'], reverse=True)
    
    # Process weekly data
    weekly_trend = []
    for week, stats in sorted(weekly_data.items()):
        total_events = stats['new'] + stats['removed'] + stats['changed']
        week_vol_rate = round(total_events / max(stats['total_tasks'], 1) * 100, 1)
        weekly_trend.append({
            'week': week,
            'volatility_rate': week_vol_rate,
            'new': stats['new'],
            'removed': stats['removed'],
            'changed': stats['changed'],
            'total_tasks': stats['total_tasks']
        })
    
    return jsonify({
        'analysis_period': {
            'days': days,
            'snapshots_analyzed': len(filtered_snapshots),
            'comparisons': total_comparisons,
            'asset_filter': asset_filter or 'All Assets'
        },
        'summary': {
            'volatility_rate': volatility_rate,
            'volatility_label': 'High' if volatility_rate > 30 else ('Medium' if volatility_rate > 15 else 'Low'),
            'avg_shift_days': avg_shift_days,
            'cancellation_rate': cancellation_rate,
            'new_route_rate': new_route_rate,
            'total_new': total_new,
            'total_removed': total_removed,
            'total_changed': total_changed,
            'total_unchanged': total_unchanged,
            'total_date_shifts': total_date_shifts
        },
        'by_asset': asset_rankings[:10],  # Top 10 most volatile assets
        'weekly_trend': weekly_trend,
        'comparisons': comparisons[-10:]  # Last 10 comparisons
    })


@app.route('/api/snapshots', methods=['POST'])
def create_snapshot():
    """Manually create a snapshot from the current task data.
    
    Use this to establish a baseline for planned vs actual tracking.
    """
    try:
        with open(DATA_DIR / 'tasks.json', 'r') as f:
            current_data = json.load(f)
    except FileNotFoundError:
        return jsonify({'error': 'No task data to snapshot'}), 400
    
    timestamp = datetime.now()
    snapshot_filename = f"tasks_{timestamp.strftime('%Y%m%d_%H%M%S')}.json"
    
    snapshot_data = {
        **current_data,
        'snapshot_id': snapshot_filename,
        'snapshot_date': timestamp.strftime('%Y-%m-%d'),
        'snapshot_time': timestamp.strftime('%H:%M:%S'),
        'imported_at': timestamp.isoformat(),
        'snapshot_type': 'manual'
    }
    
    with open(SNAPSHOTS_DIR / snapshot_filename, 'w') as f:
        json.dump(snapshot_data, f, indent=2)
    
    return jsonify({
        'success': True,
        'snapshot_id': snapshot_filename,
        'date': timestamp.strftime('%Y-%m-%d'),
        'time': timestamp.strftime('%H:%M:%S'),
        'task_count': len(current_data.get('tasks', []))
    })


@app.route('/api/snapshots/<snapshot_id>', methods=['DELETE'])
def delete_snapshot(snapshot_id):
    """Delete a specific snapshot."""
    snapshot_path = SNAPSHOTS_DIR / snapshot_id
    if not snapshot_path.exists():
        return jsonify({'error': 'Snapshot not found'}), 404
    try:
        snapshot_path.unlink()
        return jsonify({'success': True})
    except IOError as e:
        return jsonify({'error': str(e)}), 500


# ============== File Upload ==============

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Allowed: xlsx, xls, json'}), 400
    
    filename = secure_filename(file.filename)
    filepath = UPLOAD_FOLDER / filename
    file.save(filepath)
    
    try:
        if filename.endswith('.json'):
            # Direct JSON upload
            result = process_json_upload(filepath)
        else:
            # Excel upload
            if not EXCEL_SUPPORT:
                return jsonify({
                    'error': 'Excel support not available. Install openpyxl: pip install openpyxl'
                }), 400
            result = process_excel_upload(filepath, filename)
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500


def process_json_upload(filepath):
    """Process a direct JSON file upload."""
    with open(filepath, 'r') as f:
        data = json.load(f)
    
    # Detect which type of data this is and save to appropriate file
    if 'tasks' in data:
        with open(DATA_DIR / 'tasks.json', 'w') as f:
            json.dump(data, f, indent=2)
        return {'success': True, 'type': 'tasks', 'count': len(data.get('tasks', []))}
    
    elif 'records' in data:
        with open(DATA_DIR / 'spot-hire.json', 'w') as f:
            json.dump(data, f, indent=2)
        return {'success': True, 'type': 'spot-hire', 'count': len(data.get('records', []))}
    
    elif 'conflicts' in data:
        with open(DATA_DIR / 'conflicts.json', 'w') as f:
            json.dump(data, f, indent=2)
        return {'success': True, 'type': 'conflicts', 'count': len(data.get('conflicts', []))}
    
    elif 'asset_capacities' in data:
        with open(DATA_DIR / 'asset-capacity.json', 'w') as f:
            json.dump(data, f, indent=2)
        return {'success': True, 'type': 'asset-capacity', 'count': len(data.get('asset_capacities', []))}
    
    return {'error': 'Unknown JSON format'}, 400


def process_excel_upload(filepath, filename):
    """Process an Excel workbook upload."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = {'success': True, 'sheets_processed': [], 'tasks': 0, 'spot_hire': 0}
    
    # Look for known sheet patterns
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Detect sheet type by headers or name
        headers = [safe_string(cell.value).lower() for cell in sheet[1]]
        
        if is_tasks_sheet(sheet_name, headers):
            tasks = parse_tasks_sheet(sheet, filename, sheet_name)
            if tasks:
                save_tasks_data(tasks, filename, sheet_name)
                results['sheets_processed'].append(sheet_name)
                results['tasks'] = len(tasks)
        
        elif is_spot_hire_sheet(sheet_name, headers):
            records = parse_spot_hire_sheet(sheet, filename, sheet_name)
            if records:
                save_spot_hire_data(records, filename, sheet_name)
                results['sheets_processed'].append(sheet_name)
                results['spot_hire'] = len(records)
    
    wb.close()
    return results


# Phase color mapping - distinct colors for each category
PHASE_COLORS = {
    'EV Run': '#b79ac7',           # Light Purple
    'Top Hole': '#00b4d8',         # Cyan
    'Deepening': '#ff6b35',        # Orange
    'Completion': '#7cb518',       # Green
    'Demob': '#6c757d',            # Gray
    'PA': '#9d4edd',               # Purple
    'TA': '#3a86ff',               # Blue
    'Dedicated Run': '#ffd60a',    # Yellow/Gold
    'Frac Spot Hire': '#e76f51',   # Coral/Red
    'Other': '#f72585',            # Hot Pink (distinct)
    # Aliases for backwards compatibility
    'Abandonment': '#9d4edd',      # Same as PA
    'P&A': '#9d4edd',              # Same as PA
    'Turnaround': '#3a86ff',       # Same as TA
    'Dedicated OSV': '#ffd60a',    # Same as Dedicated Run
    'Pilot Hole': '#00b4d8',       # Same as Top Hole
    'Idle': '#adb5bd',             # Light gray
}


def infer_phase_from_activity(activity):
    """Infer the phase/type from activity text using keyword matching."""
    if not activity:
        return 'Other'
    
    text = activity.lower()
    
    # Check patterns in order of specificity
    if 'ev run' in text or 'ev ' in text:
        return 'EV Run'
    if 'frac' in text:
        return 'Frac Spot Hire'
    if 'demob' in text or 'demobilization' in text:
        return 'Demob'
    if 'dedicated' in text or 'dedicated run' in text:
        return 'Dedicated Run'
    if ' ta' in text or text.startswith('ta ') or 'turnaround' in text or 'turn around' in text:
        return 'TA'
    if 'p&a' in text or ' pa' in text or text.endswith(' pa') or text.endswith('pa') or 'abandon' in text:
        return 'PA'
    if 'completion' in text or 'complete' in text:
        return 'Completion'
    if 'top hole' in text or 'tophole' in text or 'pilot hole' in text or 'pilot' in text:
        return 'Top Hole'
    if 'deepen' in text or 'sidetrack' in text or ('drill' in text and 'pilot' not in text and 'top' not in text):
        return 'Deepening'
    if 'idle' in text:
        return 'Other'
    
    return 'Other'


def is_tasks_sheet(name, headers):
    """Check if sheet contains task data."""
    name_lower = name.lower()
    task_indicators = ['route', 'demand', 'task', 'schedule', 'osv']
    header_indicators = ['asset', 'activity', 'start_date', 'offshore', 'project']
    
    if any(ind in name_lower for ind in task_indicators):
        return True
    if sum(1 for h in header_indicators if h in ' '.join(headers)) >= 3:
        return True
    return False


def is_spot_hire_sheet(name, headers):
    """Check if sheet contains spot hire data."""
    name_lower = name.lower()
    spot_indicators = ['spot', 'hire', 'charter']
    header_indicators = ['phase', 'area', 'start', 'end']
    
    if any(ind in name_lower for ind in spot_indicators):
        return True
    if sum(1 for h in header_indicators if h in ' '.join(headers)) >= 2:
        return True
    return False


def parse_tasks_sheet(sheet, filename, sheet_name):
    """Parse a tasks sheet from Excel."""
    headers = [safe_string(cell.value).lower().replace(' ', '_') for cell in sheet[1]]
    print(f"DEBUG: Sheet '{sheet_name}' headers: {headers}")
    
    # Map common column variations (including Asset Activity Tracker format)
    column_map = {
        'asset': find_column(headers, ['asset', 'platform', 'rig']),
        'coordinator': find_column(headers, ['coordinator', 'planner', 'owner']),
        'vessel': find_column(headers, ['vessel', 'boat', 'osv']),
        'project': find_column(headers, ['project', 'well', 'campaign']),
        'activity': find_column(headers, ['activity', 'activity_name', 'description', 'task', 'scope', 'cargo', 'load', 'run', 'shipment']),
        'status': find_column(headers, ['status', 'state']),
        'start_date': find_column(headers, ['start_date', 'start', 'sail_date', 'base_delivery_date', 'delivery_date', 'sail']),
        'offshore_start': find_column(headers, ['offshore_start', 'arrive', 'arrival', 'offshore_req_date', 'offshore_req', 'req_date']),
        'offshore_end': find_column(headers, ['offshore_end', 'depart', 'departure', 'offloading_complete_date', 'offloading_complete', 'complete_date']),
        'return_end': find_column(headers, ['return_end', 'return', 'back', 'back_to_port', 'port']),
        'duration_hours': find_column(headers, ['duration_hours', 'duration', 'hours', 'estimated_duration', 'hrs']),
        'transit_hours': find_column(headers, ['transit_hours', 'transit', 'steam', 'transit_time', 'transit_time_back']),
    }
    print(f"DEBUG: Column mappings: {column_map}")
    
    tasks = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        task = {
            'id': generate_id(),
            'asset': get_cell_value(row, column_map['asset']),
            'coordinator': get_cell_value(row, column_map['coordinator']),
            'vessel': get_cell_value(row, column_map['vessel'], 'Route Demand'),
            'project': get_cell_value(row, column_map['project']),
            'activity': get_cell_value(row, column_map['activity']),
            'status': get_cell_value(row, column_map['status'], 'Planned'),
            'start_date': parse_date(get_cell_value(row, column_map['start_date'])),
            'offshore_start': parse_date(get_cell_value(row, column_map['offshore_start'])),
            'offshore_end': parse_date(get_cell_value(row, column_map['offshore_end'])),
            'return_end': parse_date(get_cell_value(row, column_map['return_end'])),
            'duration_hours': safe_float(get_cell_value(row, column_map['duration_hours']), 24.0),
            'transit_hours': safe_float(get_cell_value(row, column_map['transit_hours']), 18.0),
            'source': 'workbook',
            'sheet_row': row_idx
        }
        
        # Skip rows without essential data
        if task['asset'] and (task['activity'] or task['project']):
            tasks.append(task)
    
    return tasks


def find_header_row(sheet, required_headers, max_rows=20):
    """Find the row containing headers by searching for required column names."""
    for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
        row_values = [safe_string(cell.value).lower() for cell in sheet[row_idx]]
        row_text = ' '.join(row_values)
        matches = sum(1 for h in required_headers if h in row_text)
        if matches >= 2:  # At least 2 required headers found
            return row_idx, row_values
    return 1, [safe_string(cell.value).lower() for cell in sheet[1]]  # Default to row 1


def parse_spot_hire_sheet(sheet, filename, sheet_name):
    """Parse a spot hire sheet from Excel."""
    # Find the header row (may not be row 1)
    header_row, raw_headers = find_header_row(sheet, ['asset', 'start', 'end', 'activity'])
    headers = [h.replace(' ', '_') for h in raw_headers]
    print(f"DEBUG: Spot hire sheet '{sheet_name}' header row: {header_row}, headers: {headers[:10]}")
    
    column_map = {
        'asset': find_column(headers, ['asset', 'platform', 'rig']),
        'vessel_count': find_column(headers, ['vessel_count', 'vessels', 'osv_count']),
        'area': find_column(headers, ['area', 'region', 'location']),
        'activity': find_column(headers, ['activity', 'description', 'scope', 'campaign', 'well_scope']),
        'phase': find_column(headers, ['phase', 'type', 'category']),
        'start_date': find_column(headers, ['start_date', 'start', 'from']),
        'end_date': find_column(headers, ['end_date', 'end', 'to']),
        'status': find_column(headers, ['status', 'state']),
        'notes': find_column(headers, ['notes', 'comments', 'remarks']),
    }
    print(f"DEBUG: Spot hire column mappings: {column_map}")
    
    records = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(row):
            continue
        
        asset_val = get_cell_value(row, column_map['asset'])
        if not asset_val:
            continue
        
        activity = get_cell_value(row, column_map['activity'])
        explicit_phase = get_cell_value(row, column_map['phase']) if column_map['phase'] is not None else ''
        
        # Use explicit phase if available, otherwise infer from activity
        phase = explicit_phase if explicit_phase else infer_phase_from_activity(activity)
        color = PHASE_COLORS.get(phase, PHASE_COLORS['Other'])
            
        # Parse vessel count (default to 1 if not specified)
        vessel_count_val = get_cell_value(row, column_map['vessel_count'], '1')
        try:
            vessel_count = max(1, min(10, int(float(vessel_count_val))))
        except (ValueError, TypeError):
            vessel_count = 1
        
        record = {
            'id': generate_id(),
            'asset': asset_val.strip(),
            'display_asset': asset_val.strip(),
            'vessel_count': vessel_count,
            'area': get_cell_value(row, column_map['area']),
            'activity': activity,
            'phase': phase,
            'color': color,
            'start_date': parse_date(get_cell_value(row, column_map['start_date'])),
            'end_date': parse_date(get_cell_value(row, column_map['end_date'])),
            'status': get_cell_value(row, column_map['status'], 'Planned'),
            'notes': get_cell_value(row, column_map['notes']),
            'source': 'workbook',
            'sheet_row': row_idx
        }
        
        if record['asset'] and record['start_date']:
            records.append(record)
    
    return records


def find_column(headers, variations):
    """Find column index matching any of the variations. Uses exact match priority."""
    # First try exact match
    for i, header in enumerate(headers):
        for var in variations:
            if header == var:
                return i
    # Then try substring match
    for i, header in enumerate(headers):
        for var in variations:
            if var in header and var != 'count':  # Avoid matching 'count' in 'vessel_count'
                return i
    return None


def get_cell_value(row, col_idx, default=''):
    """Get cell value safely."""
    if col_idx is None or col_idx >= len(row):
        return default
    value = row[col_idx]
    return safe_string(value) if value is not None else default


def save_tasks_data(tasks, filename, sheet_name):
    """Save parsed tasks to JSON file and create a snapshot for history tracking."""
    timestamp = datetime.now()
    data = {
        'tasks': tasks,
        'source': f'workbook:{filename}:{sheet_name}',
        'buffer_hours': 24,
        'imported_at': timestamp.isoformat()
    }
    
    # Save current tasks
    with open(DATA_DIR / 'tasks.json', 'w') as f:
        json.dump(data, f, indent=2)
    
    # Save snapshot for planned vs actual tracking
    snapshot_filename = f"tasks_{timestamp.strftime('%Y%m%d_%H%M%S')}.json"
    snapshot_data = {
        **data,
        'snapshot_id': snapshot_filename,
        'snapshot_date': timestamp.strftime('%Y-%m-%d'),
        'snapshot_time': timestamp.strftime('%H:%M:%S')
    }
    with open(SNAPSHOTS_DIR / snapshot_filename, 'w') as f:
        json.dump(snapshot_data, f, indent=2)


# Standard phases that should always appear in the legend
STANDARD_PHASES = [
    'EV Run', 'Top Hole', 'Deepening', 'Completion', 'Demob',
    'PA', 'TA', 'Dedicated Run', 'Frac Spot Hire', 'Other'
]


def save_spot_hire_data(records, filename, sheet_name):
    """Save parsed spot hire records to JSON file."""
    # Always include all standard phases in color mapping for consistent legend
    phase_colors = {phase: PHASE_COLORS.get(phase, PHASE_COLORS['Other']) for phase in STANDARD_PHASES}

    # Also add any additional phases from records
    for record in records:
        phase = record.get('phase')
        if phase and phase not in phase_colors:
            phase_colors[phase] = PHASE_COLORS.get(phase, PHASE_COLORS['Other'])

    data = {
        'source': f'workbook:{filename}:{sheet_name}',
        'sheet_name': sheet_name,
        'records': records,
        'phase_colors': phase_colors,
        'imported_at': datetime.now().isoformat()
    }
    with open(DATA_DIR / 'spot-hire.json', 'w') as f:
        json.dump(data, f, indent=2)

    # Auto-update asset capacity from spot hire data
    update_asset_capacity_from_spot_hire(records)


def update_asset_capacity_from_spot_hire(records):
    """Update asset-capacity.json from spot hire records.

    Only includes records with status='Planned' that haven't ended yet.
    """
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    asset_capacities = []

    for record in records:
        # Only include 'Planned' status
        status = record.get('status', '').strip()
        if status.lower() != 'planned':
            continue

        # Parse dates
        start_str = record.get('start_date', '')
        end_str = record.get('end_date', '')

        try:
            if 'T' in str(start_str):
                start_date = datetime.fromisoformat(str(start_str).replace('Z', '+00:00')).replace(tzinfo=None)
            else:
                start_date = datetime.strptime(str(start_str).split()[0], '%Y-%m-%d')

            if 'T' in str(end_str):
                end_date = datetime.fromisoformat(str(end_str).replace('Z', '+00:00')).replace(tzinfo=None)
            else:
                end_date = datetime.strptime(str(end_str).split()[0], '%Y-%m-%d')
        except (ValueError, TypeError):
            continue

        # Skip activities that have already ended
        if end_date < today:
            continue

        # Build notes from activity/phase
        activity = record.get('activity', '')
        phase = record.get('phase', '')
        if activity and phase and phase.lower() not in activity.lower():
            notes = f"{activity} {phase}".strip()
        else:
            notes = activity or phase or ''

        asset_capacities.append({
            "asset": record.get('asset', ''),
            "vessel_count": int(record.get('vessel_count', 1) or 1),
            "notes": notes,
            "date_from": start_date.strftime("%Y-%m-%d"),
            "date_to": end_date.strftime("%Y-%m-%d")
        })

    # Sort by start date
    asset_capacities.sort(key=lambda x: x["date_from"])

    # Calculate monthly capacity entries
    monthly_capacity = calculate_monthly_capacity(asset_capacities)

    # Save to asset-capacity.json
    capacity_data = {
        "asset_capacities": asset_capacities,
        "monthly_capacity_entries": monthly_capacity,
        "updated_at": datetime.now().isoformat(),
        "source": "auto-sync from Spot Hire Planner"
    }

    with open(DATA_DIR / 'asset-capacity.json', 'w') as f:
        json.dump(capacity_data, f, indent=2)

    print(f"Auto-updated asset capacity: {len(asset_capacities)} entries, {len(monthly_capacity)} months")


def calculate_monthly_capacity(asset_capacities):
    """Calculate maximum vessel count needed per month."""
    if not asset_capacities:
        return []

    today = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_date = today + timedelta(days=6 * 31)  # 6 months ahead

    monthly_entries = []
    current_month = today

    while current_month < end_date:
        # Get first and last day of month
        month_start = current_month.replace(day=1)
        if current_month.month == 12:
            next_month = current_month.replace(year=current_month.year + 1, month=1, day=1)
        else:
            next_month = current_month.replace(month=current_month.month + 1, day=1)
        month_end = next_month - timedelta(days=1)

        # Calculate max vessels needed on any day in this month
        max_vessels = 0
        current_day = month_start

        while current_day <= month_end:
            daily_vessels = 0
            for cap in asset_capacities:
                try:
                    cap_start = datetime.strptime(cap["date_from"], "%Y-%m-%d")
                    cap_end = datetime.strptime(cap["date_to"], "%Y-%m-%d")
                    if cap_start <= current_day <= cap_end:
                        daily_vessels += cap["vessel_count"]
                except (ValueError, KeyError):
                    continue
            max_vessels = max(max_vessels, daily_vessels)
            current_day += timedelta(days=1)

        if max_vessels > 0:
            monthly_entries.append({
                "capacity_text": str(max_vessels),
                "date_from": month_start.strftime("%Y-%m-%d"),
                "date_to": month_end.strftime("%Y-%m-%d")
            })

        current_month = next_month

    return monthly_entries


# ============== Main ==============

if __name__ == '__main__':
    print("\n" + "=" * 50)
    print("OSV Demand Scheduler Local Server")
    print("=" * 50)
    print(f"Excel support: {'Yes' if EXCEL_SUPPORT else 'No (install openpyxl)'}")
    print(f"Data directory: {DATA_DIR}")
    print(f"Upload directory: {UPLOAD_FOLDER}")
    print("\nStarting server at http://127.0.0.1:8000")
    print("Press Ctrl+C to stop\n")
    
    app.run(host='127.0.0.1', port=8000, debug=True)
