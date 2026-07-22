"""
Combined OSV Scheduler Update Workbook Generator
Creates one Excel workbook with both update tabs:
- OSV Demand Tracker
- Spot Hire Update

The existing upload endpoint already processes multiple recognized sheets in one workbook.
"""
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    raise SystemExit(1)

DATA_DIR = Path("data")
OUTPUT_FILE = f"OSV_Scheduler_Combined_Update_{datetime.now().strftime('%Y%m%d')}.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_COLORS = {
    "Complete": "7CB518",
    "Completed": "7CB518",
    "In Progress": "FFD60A",
    "Planned": "00B4D8",
    "On Hold": "6C757D",
    "Cancelled": "F72585",
}

PHASE_COLORS = {
    "EV Run": "B79AC7",
    "Top Hole": "00B4D8",
    "Deepening": "FF6B35",
    "Completion": "7CB518",
    "Demob": "6C757D",
    "PA": "9D4EDD",
    "TA": "3A86FF",
    "Dedicated Run": "FFD60A",
    "Frac Spot Hire": "E76F51",
    "Other": "F72585",
}


def load_json(path, fallback):
    if not path.exists():
        return fallback
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def style_header(ws, headers, widths):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def format_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value).split()[0].split("T")[0], "%Y-%m-%d")
    except (ValueError, TypeError):
        return value


def format_datetime(value):
    if not value:
        return None
    raw = str(value).replace("T", " ").split(".")[0]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return value


def set_cell(ws, row, column, value, wrap=False, number_format=None):
    cell = ws.cell(row=row, column=column, value=value)
    cell.border = THIN_BORDER
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    if number_format and isinstance(value, datetime):
        cell.number_format = number_format
    return cell


def build_route_sheet(wb, tasks_data):
    ws = wb.active
    ws.title = "OSV Demand Tracker"
    headers = [
        "Status",
        "Asset",
        "Project",
        "Activity",
        "Loading Time",
        "Base Delivery Date",
        "Offshore Req Date",
        "Offloading Complete Date",
        "Estimated Duration (hrs.)",
        "Back to Port",
        "Transit Time Back to Port",
    ]
    widths = [12, 15, 25, 50, 16, 16, 18, 20, 20, 18, 22]
    style_header(ws, headers, widths)

    tasks = tasks_data.get("tasks", [])
    for row_idx, task in enumerate(tasks, 2):
        status = task.get("status", "Planned")
        status_cell = set_cell(ws, row_idx, 1, status)
        if status in STATUS_COLORS:
            status_cell.fill = PatternFill(start_color=STATUS_COLORS[status], end_color=STATUS_COLORS[status], fill_type="solid")
        set_cell(ws, row_idx, 2, task.get("asset", ""))
        set_cell(ws, row_idx, 3, task.get("project", ""))
        set_cell(ws, row_idx, 4, task.get("activity", ""), wrap=True)
        set_cell(ws, row_idx, 5, format_datetime(task.get("loading_time", "")), number_format="YYYY-MM-DD HH:MM")
        set_cell(ws, row_idx, 6, format_date(task.get("start_date", "")), number_format="YYYY-MM-DD")
        set_cell(ws, row_idx, 7, format_datetime(task.get("offshore_start", "")), number_format="YYYY-MM-DD HH:MM")
        set_cell(ws, row_idx, 8, format_datetime(task.get("offshore_end", "")), number_format="YYYY-MM-DD HH:MM")
        set_cell(ws, row_idx, 9, task.get("duration_hours", 24))
        set_cell(ws, row_idx, 10, format_datetime(task.get("return_end", "")), number_format="YYYY-MM-DD HH:MM")
        set_cell(ws, row_idx, 11, task.get("transit_hours", 18))

    first_blank = len(tasks) + 2
    for row_idx in range(first_blank, first_blank + 15):
        for col in range(1, 12):
            set_cell(ws, row_idx, col, "")
        ws.cell(row=row_idx, column=1, value="Planned")
        ws.cell(row=row_idx, column=9, value=24)
        ws.cell(row=row_idx, column=11, value=18)

    return len(tasks)


def build_spot_hire_sheet(wb, spot_data):
    ws = wb.create_sheet("Spot Hire Update")
    headers = ["Asset", "Vessel Count", "Area", "Activity", "Phase", "Start Date", "End Date", "Status", "Notes"]
    widths = [20, 12, 15, 35, 16, 13, 13, 14, 42]
    style_header(ws, headers, widths)

    records = spot_data.get("records", [])
    for row_idx, record in enumerate(records, 2):
        set_cell(ws, row_idx, 1, record.get("asset", ""))
        set_cell(ws, row_idx, 2, record.get("vessel_count", 1))
        set_cell(ws, row_idx, 3, record.get("area", ""))
        set_cell(ws, row_idx, 4, record.get("activity", ""), wrap=True)
        phase = record.get("phase", "")
        phase_cell = set_cell(ws, row_idx, 5, phase)
        if phase in PHASE_COLORS:
            phase_cell.fill = PatternFill(start_color=PHASE_COLORS[phase], end_color=PHASE_COLORS[phase], fill_type="solid")
        set_cell(ws, row_idx, 6, format_date(record.get("start_date", "")), number_format="YYYY-MM-DD")
        set_cell(ws, row_idx, 7, format_date(record.get("end_date", "")), number_format="YYYY-MM-DD")
        set_cell(ws, row_idx, 8, record.get("status", "Planned"))
        set_cell(ws, row_idx, 9, record.get("notes", ""), wrap=True)

    first_blank = len(records) + 2
    for row_idx in range(first_blank, first_blank + 10):
        for col in range(1, 10):
            set_cell(ws, row_idx, col, "")
        ws.cell(row=row_idx, column=2, value=1)
        ws.cell(row=row_idx, column=8, value="Planned")

    return len(records)


def build_reference_sheet(wb):
    ws = wb.create_sheet("Reference")
    ws["A1"] = "Workbook Instructions"
    ws["A1"].font = Font(bold=True, size=14)
    instructions = [
        "Use this one workbook to update both apps.",
        "Edit route demand rows on the 'OSV Demand Tracker' tab.",
        "Edit spot hire rows on the 'Spot Hire Update' tab.",
        "Keep header names and sheet names unchanged.",
        "Upload this workbook from the OSV Demand Scheduler Upload Workbook button.",
        "The server will process both recognized sheets in one upload.",
    ]
    for row_idx, text in enumerate(instructions, 2):
        ws.cell(row=row_idx, column=1, value=text)

    ws["C1"] = "Valid Route Statuses"
    ws["C1"].font = Font(bold=True)
    for row_idx, status in enumerate(STATUS_COLORS.keys(), 2):
        cell = ws.cell(row=row_idx, column=3, value=status)
        cell.fill = PatternFill(start_color=STATUS_COLORS[status], end_color=STATUS_COLORS[status], fill_type="solid")

    ws["E1"] = "Valid Spot Hire Phases"
    ws["E1"].font = Font(bold=True)
    for row_idx, phase in enumerate(PHASE_COLORS.keys(), 2):
        cell = ws.cell(row=row_idx, column=5, value=phase)
        cell.fill = PatternFill(start_color=PHASE_COLORS[phase], end_color=PHASE_COLORS[phase], fill_type="solid")

    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["E"].width = 24


def main():
    tasks_data = load_json(DATA_DIR / "tasks.json", {"tasks": []})
    spot_data = load_json(DATA_DIR / "spot-hire.json", {"records": []})

    wb = Workbook()
    task_count = build_route_sheet(wb, tasks_data)
    spot_count = build_spot_hire_sheet(wb, spot_data)
    build_reference_sheet(wb)
    wb.save(OUTPUT_FILE)

    print(f"Created: {OUTPUT_FILE}")
    print(f"Route demand rows: {task_count}")
    print(f"Spot hire rows: {spot_count}")
    print("\nUpload instructions:")
    print("1. Edit both workbook tabs as needed")
    print("2. Save the workbook")
    print("3. Upload via the OSV Demand Scheduler Upload Workbook button")


if __name__ == "__main__":
    main()
