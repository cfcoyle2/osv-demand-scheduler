import openpyxl
from datetime import datetime
import json
import uuid

wb = openpyxl.load_workbook('uploads/Asset_Activity_Tracker_MASTER.xlsx', data_only=True)
ws = wb['2026 Spot Hire Update']

jan_2026 = datetime(2026, 1, 1)
jun_end = datetime(2026, 6, 30)

# Phase color mapping
phase_colors = {
    'Top Hole': '#4fc3f7',
    'Deepening': '#81c784',
    'Completion': '#ffb74d',
    'Abandonment': '#ef5350',
    'PA': '#ef5350',
    'Dedicated OSV': '#9575cd',
    'Demob': '#90a4ae',
    'Demobilization': '#90a4ae',
    'Pilot Hole': '#4db6ac',
    'Idle': '#bdbdbd',
    'EV Run': '#b79ac7',
    'Frac': '#f06292',
    'SURF': '#7986cb',
    'TAR': '#a1887f',
    'Warehouse': '#a1887f',
    'Intervention': '#ffca28',
    'PLT': '#ffca28',
    'Riser': '#26a69a',
    'Side Track': '#42a5f5',
    'Sidetrack': '#42a5f5',
    'Drill': '#66bb6a',
    'Bypass': '#ab47bc',
}

def get_phase_and_color(activity):
    activity_lower = activity.lower()
    for keyword, color in phase_colors.items():
        if keyword.lower() in activity_lower:
            return keyword, color
    return activity[:20], '#78909c'  # Default gray

records = []
row_num = 11
for row_idx in range(11, ws.max_row + 1):
    asset = ws.cell(row=row_idx, column=2).value
    area = ws.cell(row=row_idx, column=3).value
    activity = ws.cell(row=row_idx, column=4).value
    start_date = ws.cell(row=row_idx, column=5).value
    end_date = ws.cell(row=row_idx, column=6).value
    
    if not activity and not start_date:
        continue
    
    # Parse dates
    if isinstance(start_date, datetime):
        start_dt = start_date
    elif start_date:
        try:
            start_dt = datetime.fromisoformat(str(start_date).replace(' ', 'T')[:19])
        except:
            start_dt = None
    else:
        start_dt = None
        
    if isinstance(end_date, datetime):
        end_dt = end_date
    elif end_date:
        try:
            end_dt = datetime.fromisoformat(str(end_date).replace(' ', 'T')[:19])
        except:
            end_dt = None
    else:
        end_dt = None
    
    # Filter for records active during Jan-Jun 2026
    if start_dt and start_dt <= jun_end and (not end_dt or end_dt >= jan_2026):
        asset_str = str(asset).strip().replace('\n', ' ') if asset else ''
        area_str = str(area).strip() if area else ''
        activity_str = str(activity).strip().replace('\t', '').replace('\n', ' ') if activity else ''
        
        final_asset = asset_str if asset_str else area_str
        phase, color = get_phase_and_color(activity_str)
        
        record = {
            'id': uuid.uuid4().hex[:12],
            'asset': final_asset,
            'display_asset': final_asset,
            'area': area_str,
            'activity': activity_str,
            'phase': phase,
            'color': color,
            'start_date': start_dt.strftime('%Y-%m-%d %H:%M:%S') if start_dt else '',
            'end_date': end_dt.strftime('%Y-%m-%d %H:%M:%S') if end_dt else '',
            'status': 'Planned',
            'notes': '',
            'source': 'workbook',
            'sheet_row': row_num
        }
        records.append(record)
    
    row_num += 1

# Build the output
output = {
    'source': 'workbook:Asset_Activity_Tracker_MASTER.xlsx:2026 Spot Hire Update',
    'sheet_name': '2026 Spot Hire Update',
    'records': records,
    'phase_colors': phase_colors
}

# Save to file
with open('data/spot-hire.json', 'w') as f:
    json.dump(output, f, indent=2)

print(f'Updated spot-hire.json with {len(records)} records')
print()
for r in records:
    print(f"{r['asset']:<25} | {r['phase']:<15} | {r['start_date'][:10]} to {r['end_date'][:10] if r['end_date'] else 'TBD'}")
