"""
Fluid Tracking Template Generator
=================================
Creates an Excel template for tracking fluid demand during drilling and completion activities:
- Top Hole: WBM, Cement
- Deepening: SBM, WBM, Cement
- Completions: Brine, Packer Fluid, Base Oil

Based on existing OSV Demand Scheduler activities.
"""
import json
import re
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    print("openpyxl required: pip install openpyxl")
    exit(1)

# Load existing tasks to extract fluid patterns
data_file = Path('data/tasks.json')
spot_hire_file = Path('data/spot-hire.json')

existing_tasks = []
if data_file.exists():
    with open(data_file, 'r', encoding='utf-8') as f:
        existing_tasks = json.load(f).get('tasks', [])

spot_hire_data = []
if spot_hire_file.exists():
    with open(spot_hire_file, 'r', encoding='utf-8') as f:
        spot_hire_data = json.load(f).get('records', [])

# Extract unique assets and projects
assets = sorted(set(t.get('asset', '') for t in existing_tasks if t.get('asset')))
projects = sorted(set(t.get('project', '') for t in existing_tasks if t.get('project')))

# Define fluid types and their typical phases
FLUID_TYPES = {
    'WBM': {'phases': ['Top Hole', 'Deepening'], 'unit': 'bbl', 'color': '00B4D8'},
    'SBM': {'phases': ['Deepening'], 'unit': 'bbl', 'color': 'FF6B35'},
    'Brine': {'phases': ['Completion'], 'unit': 'bbl', 'color': '7CB518'},
    'Cement': {'phases': ['Top Hole', 'Deepening', 'Completion'], 'unit': 'sacks', 'color': '9D4EDD'},
    'Drill Water': {'phases': ['Top Hole', 'Deepening', 'Completion'], 'unit': 'bbl', 'color': '3A86FF'},
    'Base Oil': {'phases': ['Deepening', 'Completion'], 'unit': 'bbl', 'color': 'FFD60A'},
    'Packer Fluid': {'phases': ['Completion'], 'unit': 'bbl', 'color': 'F72585'},
}

PHASES = ['Top Hole', 'Deepening', 'Completion']
DELIVERY_TYPES = ['Delivery', 'Backload']
STATUSES = ['Planned', 'Ordered', 'In Transit', 'Delivered', 'Complete']

# Define styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Phase colors
phase_colors = {
    'Top Hole': '00B4D8',
    'Deepening': 'FF6B35',
    'Completion': '7CB518',
}


def extract_fluid_data(activity_text):
    """Extract fluid information from activity description."""
    fluids = []
    
    # Pattern: X,XXX bbl of XX ppg WBM/SBM/Brine
    bbl_pattern = r'([\d,]+)\s*(?:bbl|bbls?)\s*(?:of\s+)?([\d.]+\s*(?:ppg|#))?\s*(WBM|SBM|Brine)?'
    matches = re.findall(bbl_pattern, activity_text, re.IGNORECASE)
    for match in matches:
        volume = int(match[0].replace(',', ''))
        weight = match[1] if match[1] else ''
        fluid_type = match[2].upper() if match[2] else 'Unknown'
        fluids.append({'type': fluid_type, 'volume': volume, 'weight': weight, 'unit': 'bbl'})
    
    # Pattern: X,XXX Sacks of Gel/Cement
    sack_pattern = r'([\d,]+)\s*(?:Sacks?)\s*(?:of\s+)?(\w+)?'
    matches = re.findall(sack_pattern, activity_text, re.IGNORECASE)
    for match in matches:
        volume = int(match[0].replace(',', ''))
        material = match[1] if match[1] else 'Cement'
        fluids.append({'type': material, 'volume': volume, 'weight': '', 'unit': 'sacks'})
    
    # Check for backload
    is_backload = 'backload' in activity_text.lower()
    
    return fluids, is_backload


def create_fluid_tracking_workbook():
    """Create the fluid tracking Excel workbook."""
    wb = Workbook()
    
    # ============== Sheet 1: Fluid Requirements ==============
    ws = wb.active
    ws.title = "Fluid Requirements"
    
    # Headers
    headers = [
        'Well/Project', 'Asset', 'Phase', 'Fluid Type', 'Weight (ppg/#)',
        'Volume', 'Unit', 'Type', 'Delivery Date', 'OSV Run',
        'Status', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Column widths
    column_widths = [25, 15, 12, 12, 14, 12, 8, 10, 14, 15, 12, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add data validation
    # Phase validation
    phase_dv = DataValidation(type="list", formula1='"' + ','.join(PHASES) + '"', allow_blank=True)
    phase_dv.error = "Please select a valid phase"
    phase_dv.prompt = "Select the well phase"
    ws.add_data_validation(phase_dv)
    phase_dv.add('C2:C1000')
    
    # Fluid type validation
    fluid_dv = DataValidation(type="list", formula1='"' + ','.join(FLUID_TYPES.keys()) + '"', allow_blank=True)
    ws.add_data_validation(fluid_dv)
    fluid_dv.add('D2:D1000')
    
    # Delivery type validation
    delivery_dv = DataValidation(type="list", formula1='"' + ','.join(DELIVERY_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(delivery_dv)
    delivery_dv.add('H2:H1000')
    
    # Status validation
    status_dv = DataValidation(type="list", formula1='"' + ','.join(STATUSES) + '"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add('K2:K1000')
    
    # Asset validation (if we have assets)
    if assets:
        asset_dv = DataValidation(type="list", formula1='"' + ','.join(assets[:50]) + '"', allow_blank=True)
        ws.add_data_validation(asset_dv)
        asset_dv.add('B2:B1000')
    
    # Pre-populate with sample data extracted from existing tasks
    row = 2
    sample_entries = []
    
    for task in existing_tasks[:50]:  # Check first 50 tasks
        activity = task.get('activity', '')
        project = task.get('project', '')
        asset = task.get('asset', '')
        
        # Determine phase from project name
        phase = 'Deepening'  # Default
        if 'Top Hole' in project:
            phase = 'Top Hole'
        elif 'Completion' in project:
            phase = 'Completion'
        elif 'Deepening' in project:
            phase = 'Deepening'
        
        fluids, is_backload = extract_fluid_data(activity)
        
        for fluid in fluids:
            sample_entries.append({
                'project': project,
                'asset': asset,
                'phase': phase,
                'fluid_type': fluid['type'],
                'weight': fluid['weight'],
                'volume': fluid['volume'],
                'unit': fluid['unit'],
                'delivery_type': 'Backload' if is_backload else 'Delivery',
                'date': task.get('start_date', '')[:10] if task.get('start_date') else '',
                'osv_run': activity.split('/')[0].strip() if '/' in activity else '',
                'status': task.get('status', 'Planned'),
                'notes': ''
            })
    
    # Add sample entries (deduplicated, limit to 20)
    seen = set()
    for entry in sample_entries[:30]:
        key = (entry['project'], entry['fluid_type'], entry['volume'])
        if key in seen:
            continue
        seen.add(key)
        
        ws.cell(row=row, column=1, value=entry['project']).border = thin_border
        ws.cell(row=row, column=2, value=entry['asset']).border = thin_border
        
        phase_cell = ws.cell(row=row, column=3, value=entry['phase'])
        phase_cell.border = thin_border
        if entry['phase'] in phase_colors:
            phase_cell.fill = PatternFill(start_color=phase_colors[entry['phase']], 
                                          end_color=phase_colors[entry['phase']], fill_type='solid')
        
        ws.cell(row=row, column=4, value=entry['fluid_type']).border = thin_border
        ws.cell(row=row, column=5, value=entry['weight']).border = thin_border
        ws.cell(row=row, column=6, value=entry['volume']).border = thin_border
        ws.cell(row=row, column=7, value=entry['unit']).border = thin_border
        ws.cell(row=row, column=8, value=entry['delivery_type']).border = thin_border
        ws.cell(row=row, column=9, value=entry['date']).border = thin_border
        ws.cell(row=row, column=10, value=entry['osv_run']).border = thin_border
        ws.cell(row=row, column=11, value=entry['status']).border = thin_border
        ws.cell(row=row, column=12, value=entry['notes']).border = thin_border
        
        row += 1
        if row > 25:  # Limit sample data
            break
    
    # Add empty rows for new entries
    for _ in range(20):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    
    # ============== Sheet 2: Summary by Phase ==============
    ws_summary = wb.create_sheet("Summary by Phase")
    
    # Headers
    summary_headers = ['Phase', 'Fluid Type', 'Total Volume', 'Unit', 'Delivery Count', 'Backload Count']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    col_widths = [15, 15, 15, 10, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = width
    
    # Pre-populate summary structure
    row = 2
    for phase in PHASES:
        for fluid_type, info in FLUID_TYPES.items():
            if phase in info['phases']:
                phase_cell = ws_summary.cell(row=row, column=1, value=phase)
                phase_cell.border = thin_border
                phase_cell.fill = PatternFill(start_color=phase_colors[phase], 
                                               end_color=phase_colors[phase], fill_type='solid')
                
                ws_summary.cell(row=row, column=2, value=fluid_type).border = thin_border
                ws_summary.cell(row=row, column=3, value=0).border = thin_border
                ws_summary.cell(row=row, column=4, value=info['unit']).border = thin_border
                ws_summary.cell(row=row, column=5, value=0).border = thin_border
                ws_summary.cell(row=row, column=6, value=0).border = thin_border
                row += 1
    
    # ============== Sheet 3: Typical Requirements Reference ==============
    ws_ref = wb.create_sheet("Typical Requirements")
    
    ref_headers = ['Activity Phase', 'Fluid Type', 'Typical Volume Range', 'Unit', 'Weight Range', 'Notes']
    for col, header in enumerate(ref_headers, 1):
        cell = ws_ref.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    col_widths = [15, 15, 20, 10, 15, 45]
    for i, width in enumerate(col_widths, 1):
        ws_ref.column_dimensions[get_column_letter(i)].width = width
    
    # Typical requirements based on observed data
    typical_data = [
        ('Top Hole', 'WBM', '10,000 - 15,000', 'bbl', '16 ppg', 'For jetting, 32" section, 28" casing'),
        ('Top Hole', 'Cement', '500 - 2,000', 'sacks', '-', '28" and 22" conductor cement jobs'),
        ('Top Hole', 'Drill Water', '2,000 - 5,000', 'bbl', '-', 'General drilling operations'),
        ('Deepening', 'SBM', '4,000 - 8,000', 'bbl', '12-14 ppg', 'For 17-1/2", 12-1/4", 8-1/2" sections'),
        ('Deepening', 'WBM', '5,000 - 10,000', 'bbl', '14-16 ppg', 'Displacement, spacer'),
        ('Deepening', 'Cement', '1,000 - 3,000', 'sacks', '-', 'Casing cement jobs, plug backs'),
        ('Deepening', 'Base Oil', '500 - 2,000', 'bbl', '-', 'SBM makeup, lost circulation'),
        ('Completion', 'Brine', '8,000 - 12,000', 'bbl', '14-15 ppg', 'Well displacement, kill fluid'),
        ('Completion', 'Packer Fluid', '500 - 2,000', 'bbl', '10-12 ppg', 'Annular packer fluid'),
        ('Completion', 'Base Oil', '500 - 1,500', 'bbl', '-', 'Lubricant, equipment prep'),
        ('Completion', 'Drill Water', '1,000 - 3,000', 'bbl', '-', 'Flushing, cleaning'),
    ]
    
    row = 2
    for data in typical_data:
        phase_cell = ws_ref.cell(row=row, column=1, value=data[0])
        phase_cell.border = thin_border
        phase_cell.fill = PatternFill(start_color=phase_colors[data[0]], 
                                       end_color=phase_colors[data[0]], fill_type='solid')
        
        ws_ref.cell(row=row, column=2, value=data[1]).border = thin_border
        ws_ref.cell(row=row, column=3, value=data[2]).border = thin_border
        ws_ref.cell(row=row, column=4, value=data[3]).border = thin_border
        ws_ref.cell(row=row, column=5, value=data[4]).border = thin_border
        
        notes_cell = ws_ref.cell(row=row, column=6, value=data[5])
        notes_cell.border = thin_border
        notes_cell.alignment = wrap_align
        
        row += 1
    
    # ============== Sheet 4: Monthly Forecast ==============
    ws_forecast = wb.create_sheet("Monthly Forecast")
    
    # Create monthly columns
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    forecast_headers = ['Fluid Type', 'Unit'] + months + ['Total']
    
    for col, header in enumerate(forecast_headers, 1):
        cell = ws_forecast.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    col_widths = [15, 8] + [10] * 12 + [12]
    for i, width in enumerate(col_widths, 1):
        ws_forecast.column_dimensions[get_column_letter(i)].width = width
    
    # Add fluid type rows
    row = 2
    for fluid_type, info in FLUID_TYPES.items():
        fluid_cell = ws_forecast.cell(row=row, column=1, value=fluid_type)
        fluid_cell.border = thin_border
        fluid_cell.fill = PatternFill(start_color=info['color'], end_color=info['color'], fill_type='solid')
        
        ws_forecast.cell(row=row, column=2, value=info['unit']).border = thin_border
        
        for col in range(3, 16):
            ws_forecast.cell(row=row, column=col, value=0).border = thin_border
        
        # Total formula
        total_cell = ws_forecast.cell(row=row, column=15)
        total_cell.value = f"=SUM(C{row}:N{row})"
        total_cell.border = thin_border
        total_cell.font = Font(bold=True)
        
        row += 1
    
    # Save workbook
    output_file = Path('uploads/Fluid_Tracking_Template.xlsx')
    output_file.parent.mkdir(exist_ok=True)
    wb.save(output_file)
    
    print(f"\n{'='*60}")
    print("Fluid Tracking Template Generated")
    print(f"{'='*60}")
    print(f"Output: {output_file.absolute()}")
    print(f"\nSheets created:")
    print("  1. Fluid Requirements - Main tracking sheet")
    print("  2. Summary by Phase - Aggregated view")
    print("  3. Typical Requirements - Reference data")
    print("  4. Monthly Forecast - Planning view")
    print(f"\nAssets found: {len(assets)}")
    print(f"Sample entries populated: {min(len(sample_entries), 20)}")
    print(f"{'='*60}\n")
    
    return output_file


if __name__ == '__main__':
    output = create_fluid_tracking_workbook()
    print(f"Template saved to: {output}")
