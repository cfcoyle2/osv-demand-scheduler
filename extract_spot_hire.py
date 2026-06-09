import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('uploads/Asset_Activity_Tracker_MASTER.xlsx', data_only=True)
ws = wb['2026 Spot Hire Update']

print('=== 2026 Spot Hire Update: January - June 2026 ===')
print()

jan_2026 = datetime(2026, 1, 1)
jun_end = datetime(2026, 6, 30)

records = []
for row_idx in range(11, ws.max_row + 1):
    asset = ws.cell(row=row_idx, column=2).value
    area = ws.cell(row=row_idx, column=3).value
    activity = ws.cell(row=row_idx, column=4).value
    start_date = ws.cell(row=row_idx, column=5).value
    end_date = ws.cell(row=row_idx, column=6).value
    
    if not activity and not start_date:
        continue
    
    if isinstance(start_date, datetime):
        start_dt = start_date
    elif start_date:
        try:
            start_dt = datetime.fromisoformat(str(start_date).replace(' ', 'T')[:19])
        except:
            start_dt = None
    else:
        start_dt = None
        
    if isinstance(end_date, datetime):
        end_dt = end_date
    elif end_date:
        try:
            end_dt = datetime.fromisoformat(str(end_date).replace(' ', 'T')[:19])
        except:
            end_dt = None
    else:
        end_dt = None
    
    # Filter for records active during Jan-Jun 2026
    if start_dt and start_dt <= jun_end and (not end_dt or end_dt >= jan_2026):
        asset_str = str(asset).strip().replace('\n', ' ') if asset else ''
        area_str = str(area).strip() if area else ''
        activity_str = str(activity).strip().replace('\t', '').replace('\n', ' ')[:45] if activity else ''
        start_str = start_dt.strftime('%Y-%m-%d') if start_dt else ''
        end_str = end_dt.strftime('%Y-%m-%d') if end_dt else ''
        
        # Use area if asset is empty
        final_asset = asset_str if asset_str else area_str
        records.append({'asset': final_asset, 'activity': activity_str, 'start': start_str, 'end': end_str})

print(f"{'Asset':<22} | {'Activity':<45} | {'Start':<10} | {'End':<10}")
print("=" * 100)

for r in records:
    print(f"{r['asset']:<22} | {r['activity']:<45} | {r['start']:<10} | {r['end']:<10}")

print()
print(f'Total records: {len(records)}')
