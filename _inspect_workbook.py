import openpyxl
from openpyxl.utils import get_column_letter
import os

path = r'C:\Users\Chris.Coyle\OneDrive - Shell\VS Code\Library\01_Websites\Vessel_Logistics_Gantt\backend\data\Asset Activity Tracker_MASTER.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['2026 Spot Hire Update']

print(f'Sheet: {ws.title}')
print(f'Images: {len(ws._images)}')
for i, img in enumerate(ws._images):
    anchor = img.anchor
    # anchor can be TwoCellAnchor or OneCellAnchor
    if hasattr(anchor, '_from'):
        print(f'  Img {i}: From (col={anchor._from.col}, row={anchor._from.row}) to (col={anchor.to.col}, row={anchor.to.row})')
    else:
        print(f'  Img {i}: {anchor}')

# Inspection range
row_start, row_end = 11, 45
col_start, col_end = 7, 54 # G to BB

print('\n--- Inspection ---')
for r in range(row_start, row_end + 1):
    row_data = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    is_dedicated = any('Dedicated OSV' in str(v) for v in row_data if v)
    
    fill_info = []
    cell_info = []
    
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        
        # Color
        fill = cell.fill
        color = None
        if fill and fill.start_color and fill.start_color.index != '00000000':
            color = fill.start_color.rgb if hasattr(fill.start_color, 'rgb') else fill.start_color.index
        
        if color or val or cell.hyperlink or cell.comment:
            col_letter = get_column_letter(c)
            parts = []
            if val: parts.append(f'val={val}')
            if color: parts.append(f'color={color}')
            if cell.hyperlink: parts.append(f'link={cell.hyperlink.target}')
            if cell.comment: parts.append(f'comment={cell.comment.text}')
            cell_info.append(f'{col_letter}{r}({", ".join(parts)})')
            
            if is_dedicated:
                fill_info.append(color)

    if is_dedicated or cell_info:
        ded_str = '[Dedicated OSV]' if is_dedicated else ''
        print(f'Row {r}: {ded_str}')
        if cell_info:
            print(f'  Cells: {", ".join(cell_info[:10])} {"..." if len(cell_info) > 10 else ""}')
        if is_dedicated:
            unique_colors = set(c for c in fill_info if c)
            print(f'  Fills: {unique_colors}')

