"""
Spot Hire Update Workbook Generator
Creates an Excel template pre-filled with current data for easy updates.
"""
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    exit(1)

# Load current data
data_file = Path('data/spot-hire.json')
with open(data_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Spot Hire Update"

# Define styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Phase colors for conditional formatting reference
phase_colors = {
    'EV Run': 'B79AC7',
    'Top Hole': '00B4D8',
    'Deepening': 'FF6B35',
    'Completion': '7CB518',
    'Demob': '6C757D',
    'PA': '9D4EDD',
    'TA': '3A86FF',
    'Dedicated Run': 'FFD60A',
    'Frac Spot Hire': 'E76F51',
    'Other': 'F72585'
}

# Status options
status_options = ['Planned', 'In Progress', 'Completed', 'On Hold', 'Cancelled']

# Headers
headers = ['Asset', 'Vessel Count', 'Area', 'Activity', 'Phase', 'Start Date', 'End Date', 'Status', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Column widths
column_widths = [20, 12, 15, 35, 15, 12, 12, 12, 40]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Populate with existing records
for row_idx, record in enumerate(data.get('records', []), 2):
    ws.cell(row=row_idx, column=1, value=record.get('asset', '')).border = thin_border
    ws.cell(row=row_idx, column=2, value=record.get('vessel_count', 1)).border = thin_border
    ws.cell(row=row_idx, column=3, value=record.get('area', '')).border = thin_border
    ws.cell(row=row_idx, column=4, value=record.get('activity', '')).border = thin_border
    ws.cell(row=row_idx, column=5, value=record.get('phase', '')).border = thin_border
    
    # Parse and format dates
    start_str = record.get('start_date', '')
    end_str = record.get('end_date', '')
    try:
        start_date = datetime.strptime(start_str.split()[0], '%Y-%m-%d') if start_str else None
        ws.cell(row=row_idx, column=6, value=start_date).border = thin_border
        ws.cell(row=row_idx, column=6).number_format = 'YYYY-MM-DD'
    except:
        ws.cell(row=row_idx, column=6, value=start_str).border = thin_border
    
    try:
        end_date = datetime.strptime(end_str.split()[0], '%Y-%m-%d') if end_str else None
        ws.cell(row=row_idx, column=7, value=end_date).border = thin_border
        ws.cell(row=row_idx, column=7).number_format = 'YYYY-MM-DD'
    except:
        ws.cell(row=row_idx, column=7, value=end_str).border = thin_border
    
    ws.cell(row=row_idx, column=8, value=record.get('status', 'Planned')).border = thin_border
    ws.cell(row=row_idx, column=9, value=record.get('notes', '')).border = thin_border
    
    # Color-code by phase
    phase = record.get('phase', '')
    if phase in phase_colors:
        fill = PatternFill(start_color=phase_colors[phase], end_color=phase_colors[phase], fill_type='solid')
        ws.cell(row=row_idx, column=5).fill = fill

# Add blank rows for new entries
last_row = len(data.get('records', [])) + 2
for row_idx in range(last_row, last_row + 10):
    for col in range(1, 10):
        ws.cell(row=row_idx, column=col).border = thin_border
    # Default vessel count to 1 for new rows
    ws.cell(row=row_idx, column=2, value=1)

# Add a reference sheet with phases
ws2 = wb.create_sheet("Reference")
ws2['A1'] = "Valid Phases"
ws2['A1'].font = Font(bold=True)
for i, phase in enumerate(phase_colors.keys(), 2):
    ws2.cell(row=i, column=1, value=phase)
    ws2.cell(row=i, column=1).fill = PatternFill(start_color=phase_colors[phase], end_color=phase_colors[phase], fill_type='solid')

ws2['C1'] = "Valid Statuses"
ws2['C1'].font = Font(bold=True)
for i, status in enumerate(status_options, 2):
    ws2.cell(row=i, column=3, value=status)

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['C'].width = 15

# Save
output_file = f'Spot_Hire_Update_{datetime.now().strftime("%Y%m%d")}.xlsx'
wb.save(output_file)
print(f"Created: {output_file}")
print(f"Records: {len(data.get('records', []))}")
print("\nInstructions:")
print("1. Edit the 'Spot Hire Update' sheet")
print("2. Add new rows or modify existing ones")
print("3. Save the file")
print("4. Upload via the Spot Hire Planner web page")
