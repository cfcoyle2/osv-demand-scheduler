"""
Generate Sparta OSV Boat Plan Excel Schedule
Creates a detailed Excel workbook with Gantt-style schedule and tracking sheets
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_osv_schedule_excel():
    """Create comprehensive OSV schedule Excel workbook"""
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    create_summary_sheet(wb)
    create_detailed_schedule(wb)
    create_vessel_tracker(wb)
    create_cargo_manifest(wb)
    
    # Save workbook
    output_file = r"c:\Users\Chris.Coyle\OneDrive - Shell\VS Code\Library\02_Analysis_and_Reports\Sparta_OSV_Boat_Plan\Sparta_OSV_Schedule.xlsx"
    wb.save(output_file)
    print(f"Excel file created: {output_file}")
    return output_file

def create_summary_sheet(wb):
    """Create project summary and overview sheet"""
    ws = wb.create_sheet("Project Summary", 0)
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    title_font = Font(bold=True, size=16)
    
    # Title
    ws['A1'] = "SPARTA 20K DEEPENING & COMPLETION WELL"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "OSV BOAT PLAN - PROJECT SUMMARY"
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:F2')
    
    # Project Details
    row = 4
    ws[f'A{row}'] = "PROJECT DETAILS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
    
    row += 1
    details = [
        ["Port Location:", "Fourchon, LA"],
        ["Rig Location:", "Sparta"],
        ["Transit Time:", "36 hours (each way)"],
        ["Port Loading Time:", "72 hours"],
        ["Offloading Time:", "12-18 hours"],
        ["", ""],
        ["Deepening Phase:", "December 24, 2026 - January 24, 2027 (32 days)"],
        ["Completion Phase:", "January 1, 2027 - February 23, 2027 (54 days)"],
        ["Overlap Period:", "January 1-24, 2027 (24 days - CRITICAL)"],
        ["Total Program Duration:", "86 days"],
    ]
    
    for detail in details:
        ws[f'A{row}'] = detail[0]
        ws[f'B{row}'] = detail[1]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # OSV Fleet Summary
    row += 2
    ws[f'A{row}'] = "OSV FLEET CONFIGURATION"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
    
    row += 1
    fleet_headers = ["Vessel ID", "Primary Role", "Capacity", "Total Trips", "Status"]
    for col, header in enumerate(fleet_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    fleet_data = [
        ["OSV-1", "SLB Bulk Fluids", "8,000 bbls", 12, "Primary"],
        ["OSV-2", "Tetra Bulk Fluids", "8,000 bbls", 13, "Primary"],
        ["OSV-3", "Drill Pipe & Equipment", "5,000 sq ft deck", 8, "Primary"],
        ["OSV-4", "Heavy Cargo & Mixed", "7,000 bbls + 4,000 sq ft", 10, "Primary"],
        ["OSV-5", "Standby/Peak Support", "Flex capacity", 3, "Standby (Jan 1-24)"],
        ["FSV-1", "Weekly Consumables", "2,500 sq ft + 500 bbls", 12, "Weekly Runs"],
    ]
    
    for vessel in fleet_data:
        for col, value in enumerate(vessel, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if col == 5 and "Standby" in value:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        row += 1
    
    # Cargo Summary
    row += 2
    ws[f'A{row}'] = "CARGO SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
    
    row += 1
    cargo_headers = ["Cargo Type", "Quantity", "Supplier/Source", "Vessel Assignment"]
    for col, header in enumerate(cargo_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    cargo_data = [
        ["Bulk Fluids (SLB)", "16,350 bbls", "SLB", "OSV-1"],
        ["Bulk Fluids (Tetra)", "15,071-16,571 bbls", "Tetra", "OSV-2"],
        ["Drill Pipe", "~35,257 feet", "Stena/Workstrings", "OSV-3"],
        ["MPT Tanks", "14 x 25-bbl + 1 tote", "Tetra", "OSV-3/4"],
        ["Consumables", "Various", "Multiple", "OSV-4"],
    ]
    
    for cargo in cargo_data:
        for col, value in enumerate(cargo, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    
    # Cost Estimate
    row += 2
    ws[f'A{row}'] = "ESTIMATED PROGRAM COST"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E78")
    
    row += 1
    ws[f'A{row}'] = "4 Primary OSVs (86 days):"
    ws[f'B{row}'] = "$5.16M - $8.6M"
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Standby Vessel (24 days):"
    ws[f'B{row}'] = "$360K - $600K"
    
    row += 1
    ws[f'A{row}'] = "FSV Weekly Runs (86 days):"
    ws[f'B{row}'] = "$688K - $1.03M"
    
    row += 1
    ws[f'A{row}'] = "TOTAL PROGRAM ESTIMATE:"
    ws[f'B{row}'] = "$6.2M - $10.2M"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="C00000")
    ws[f'B{row}'].font = Font(bold=True, size=11, color="C00000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25

def create_detailed_schedule(wb):
    """Create detailed Gantt-style schedule"""
    ws = wb.create_sheet("Detailed Schedule")
    
    # Headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Trip #", "Vessel", "Activity", "Start Date", "End Date", "Duration (Days)", 
               "Load Location", "Discharge Location", "Cargo Type", "Status", "Notes"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Generate schedule data
    row = 2
    trip_counter = 1
    
    # Define vessels and their schedules
    vessels = {
        "OSV-1": {"role": "SLB Fluids", "trips": 12, "start": datetime(2026, 12, 21)},
        "OSV-2": {"role": "Tetra Fluids", "trips": 13, "start": datetime(2026, 12, 21)},
        "OSV-3": {"role": "Drill Pipe", "trips": 8, "start": datetime(2026, 12, 20)},
        "OSV-4": {"role": "Heavy Cargo", "trips": 10, "start": datetime(2026, 12, 22)},
        "OSV-5": {"role": "Standby", "trips": 3, "start": datetime(2027, 1, 5)},
        "FSV-1": {"role": "Weekly Consumables", "trips": 12, "start": datetime(2026, 12, 24), "cycle": 7},
    }
    
    # Colors for different phases
    deepening_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    overlap_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    completion_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    
    for vessel_id, vessel_info in vessels.items():
        current_date = vessel_info["start"]
        is_fsv = vessel_id == "FSV-1"
        
        for trip in range(1, vessel_info["trips"] + 1):
            # FSV has shorter loading and transit times
            load_days = 2 if is_fsv else 3
            transit_days = 2 if is_fsv else 3
            
            # Loading phase
            load_start = current_date
            load_end = load_start + timedelta(days=load_days)
            
            ws.cell(row=row, column=1, value=trip_counter)
            ws.cell(row=row, column=2, value=vessel_id)
            ws.cell(row=row, column=3, value="Loading at Port")
            ws.cell(row=row, column=4, value=load_start.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=5, value=load_end.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=6, value=load_days)
            ws.cell(row=row, column=7, value="Fourchon")
            ws.cell(row=row, column=8, value="-")
            ws.cell(row=row, column=9, value=vessel_info["role"])
            ws.cell(row=row, column=10, value="Planned")
            
            # Apply phase coloring
            if load_start < datetime(2027, 1, 1):
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = deepening_fill
            elif load_start <= datetime(2027, 1, 24):
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = overlap_fill
            else:
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = completion_fill
            
            row += 1
            
            # Transit & Offload phase
            transit_start = load_end
            transit_end = transit_start + timedelta(days=3)
            
            ws.cell(row=row, column=1, value=trip_counter)
            ws.cell(row=row, column=2, value=vessel_id)
            ws.cell(row=row, column=3, value="Transit & Offload")
            ws.cell(row=row, column=4, value=transit_start.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=5, value=transit_end.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=6, value=transit_days)
            ws.cell(row=row, column=7, value="Fourchon")
            ws.cell(row=row, column=8, value="Sparta Rig")
            ws.cell(row=row, column=9, value=vessel_info["role"])
            ws.cell(row=row, column=10, value="Planned")
            if is_fsv:
                ws.cell(row=row, column=11, value="24-28h transit + 6-8h offload (FSV)")
            else:
                ws.cell(row=row, column=11, value="36h transit + 12-18h offload")
            
            # Apply phase coloring
            if transit_start < datetime(2027, 1, 1):
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = deepening_fill
            elif transit_start <= datetime(2027, 1, 24):
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = overlap_fill
            else:
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = completion_fill
            
            row += 1
            trip_counter += 1
            
            # Update for next trip
            if is_fsv and "cycle" in vessel_info:
                # FSV runs weekly (7 days between starts)
                current_date = vessel_info["start"] + timedelta(days=7 * trip)
            else:
                # OSVs add return time
                current_date = transit_end + timedelta(days=1)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'

def create_vessel_tracker(wb):
    """Create vessel tracking sheet for daily updates"""
    ws = wb.create_sheet("Vessel Tracker")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ["Date", "Vessel ID", "Current Status", "Current Location", "ETA Next Port", 
               "Cargo Loaded", "Cargo Discharged", "Delays (Hours)", "Weather Hold", "Comments"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add template rows for tracking
    vessels = ["OSV-1", "OSV-2", "OSV-3", "OSV-4", "OSV-5", "FSV-1"]
    start_date = datetime(2026, 12, 20)
    
    row = 2
    for day in range(90):  # 90 days of tracking
        current_date = start_date + timedelta(days=day)
        for vessel in vessels:
            ws.cell(row=row, column=1, value=current_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=vessel)
            ws.cell(row=row, column=3, value="")  # Status - to be filled
            ws.cell(row=row, column=4, value="")  # Location - to be filled
            ws.cell(row=row, column=5, value="")  # ETA - to be filled
            ws.cell(row=row, column=6, value="")  # Cargo loaded
            ws.cell(row=row, column=7, value="")  # Cargo discharged
            ws.cell(row=row, column=8, value=0)   # Delays
            ws.cell(row=row, column=9, value="No") # Weather hold
            ws.cell(row=row, column=10, value="") # Comments
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 35
    
    ws.freeze_panes = 'A2'

def create_cargo_manifest(wb):
    """Create cargo manifest and checklist"""
    ws = wb.create_sheet("Cargo Manifest")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Title
    ws['A1'] = "CARGO MANIFEST & TRACKING"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Headers
    row = 3
    headers = ["Item #", "Cargo Description", "Quantity", "Unit", "Supplier", 
               "Required Date", "Vessel Assigned", "Status"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Cargo items
    row += 1
    cargo_items = [
        [1, "Rheguard SBM (16.3 PPG)", "6,000", "bbls", "SLB", "2026-12-21", "OSV-1", "Planned"],
        [2, "DIPRO (16.3 PPG)", "7,000", "bbls", "SLB", "2026-12-21", "OSV-1", "Planned"],
        [3, "DIPRO PRE-MIX (16.3 PPG)", "3,000", "bbls", "SLB", "2026-12-21", "OSV-1", "Planned"],
        [4, "DIPRO DIF Solids Free (16.5 PPG)", "350", "bbls", "SLB", "2026-12-21", "OSV-1", "Planned"],
        [5, "CaBr2/ZnBr2 (16.5-17.1#)", "8,500-10,000", "bbls", "Tetra", "2026-12-21", "OSV-2", "Planned"],
        [6, "CaBr2 Spacer Brine (14.2#)", "700", "bbls", "Tetra", "2026-12-21", "OSV-2", "Planned"],
        [7, "CaBr2/ZnBr2 Spike (19.2#)", "276", "bbls", "Tetra", "2027-01-05", "OSV-2", "Planned"],
        [8, "NaBr Packer Fluid (12.4#)", "5,500", "bbls", "Tetra", "2027-01-15", "OSV-2", "Planned"],
        [9, "CaBr2 in 25-bbl MPTs (14.2#)", "46", "bbls (2 MPTs)", "Tetra", "2027-01-15", "OSV-3", "Planned"],
        [10, "NaBr in 25-bbl MPTs (12.4#)", "46", "bbls (2 MPTs)", "Tetra", "2027-01-20", "OSV-3", "Planned"],
        [11, "NaBr in 350-gal tote (12.4#)", "3", "bbls (1 tote)", "Tetra", "2027-01-20", "OSV-3", "Planned"],
        [12, "CaBr2/ZnBr2 in 25-bbl MPTs", "276", "bbls (12 MPTs)", "Tetra", "2027-01-10", "OSV-3", "Planned"],
        [13, "9⅝\" Drill Pipe V150", "4,357", "feet", "Stena/Workstrings", "2026-12-20", "OSV-3", "Planned"],
        [14, "5½\" Drill Pipe V150 Delta576", "7,900", "feet", "Stena/Workstrings", "2026-12-20", "OSV-3", "Planned"],
        [15, "4½\" Drill Pipe V150 Delta425 (20 lb/ft)", "5,000", "feet", "Stena/Workstrings", "2026-12-23", "OSV-3", "Planned"],
        [16, "4½\" Drill Pipe V150 Delta425 (16.6 lb/ft)", "18,000", "feet", "Stena/Workstrings", "2026-12-27", "OSV-3", "Planned"],
    ]
    
    for item in cargo_items:
        for col, value in enumerate(item, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if col == 8:  # Status column
                if value == "Planned":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    ws.freeze_panes = 'A4'

if __name__ == "__main__":
    print("Generating Sparta OSV Schedule Excel file...")
    output_file = create_osv_schedule_excel()
    print("Done!")
