"""Loads the 2026 Spot Hire Update worksheet into editable schedule records."""
from __future__ import annotations

import hashlib
import re
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from config import SPOT_HIRE_SHEET_NAME, WORKBOOK_PATH

TIMELINE_START_COL = 7  # G
TIMELINE_END_COL = 54  # BB
DATA_START_ROW = 11
DATA_END_ROW = 45
YEAR = 2026

PHASE_COLORS = {
    "FFFF0000": {"phase": "Abandonment", "color": "#e03131"},
    "FF92D050": {"phase": "Completion", "color": "#4f9d45"},
    "FFFFC000": {"phase": "Deepening", "color": "#f0a800"},
    "FFCAB9CA": {"phase": "Dedicated OSV", "color": "#b79ac7"},
    "FF00B0F0": {"phase": "Top Hole", "color": "#16858c"},
    "FF7030A0": {"phase": "Demob", "color": "#7a4fb0"},
    "INDIVIDUAL_RUN": {"phase": "Individual Run", "color": "#00a3e0"},
    "3": {"phase": "Top Hole", "color": "#16858c"},
    "1": {"phase": "Demob", "color": "#7a4fb0"},
}

DEFAULT_PHASE_COLOR = "#16858c"
YEAR_START = datetime(YEAR, 1, 1)
YEAR_END = datetime(YEAR, 12, 31)
SPOT_HIRE_DRAWING_PATH = "xl/drawings/drawing2.xml"
BLUE_VESSEL_REL_ID = "rId3"
ASSET_COUNT_RE = re.compile(r"\s*\((\d+)\)\s*$")


def _stable_id(*parts: Any) -> str:
    raw = "|".join("" if part is None else str(part) for part in parts)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]


def _fill_rgb(cell) -> str | None:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if color.type == "rgb" and isinstance(color.rgb, str) and color.rgb != "00000000":
        return color.rgb.upper()
    if color.type == "indexed" and color.indexed is not None:
        return str(color.indexed)
    return None


def _cell_value(ws, row: int, col: int) -> Any:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _to_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _clean_text(value: str) -> str:
    return " ".join(str(value).split())


def _asset_display_and_count(asset: str) -> tuple[str, int | None]:
    clean_asset = _clean_text(asset)
    match = ASSET_COUNT_RE.search(clean_asset)
    if not match:
        return clean_asset, None
    return clean_asset[:match.start()].strip(), int(match.group(1))


def _record_identity(asset: str) -> dict:
    display_asset, source_asset_count = _asset_display_and_count(asset)
    data = {"display_asset": display_asset}
    if source_asset_count is not None:
        data["source_asset_count"] = source_asset_count
    return data


def _clip_to_year(start: datetime, end: datetime) -> tuple[datetime, datetime] | None:
    if end < YEAR_START or start > YEAR_END:
        return None
    return max(start, YEAR_START), min(end, YEAR_END)


def _month_for_column(ws, col: int) -> int | None:
    month_value = _cell_value(ws, 8, col)
    if month_value is None:
        return None
    month_name = str(month_value).strip()
    for month in range(1, 13):
        if datetime(YEAR, month, 1).strftime("%B").lower() == month_name.lower():
            return month
    return None


def _timeline_dates(ws) -> dict[int, datetime]:
    dates: dict[int, datetime] = {}
    for col in range(TIMELINE_START_COL, TIMELINE_END_COL + 1):
        month = _month_for_column(ws, col)
        day = ws.cell(9, col).value
        if month is None or day is None:
            continue
        try:
            dates[col] = datetime(YEAR, month, int(day))
        except (TypeError, ValueError):
            continue
    return dates


def _segment_end(col: int, dates: dict[int, datetime]) -> datetime:
    next_cols = [candidate for candidate in dates if candidate > col]
    if next_cols:
        return dates[min(next_cols)] - timedelta(days=1)
    return dates[col] + timedelta(days=6)


def _phase_from_color(rgb: str | None) -> tuple[str, str]:
    if rgb in PHASE_COLORS:
        phase = PHASE_COLORS[rgb]
        return phase["phase"], phase["color"]
    return "Other", DEFAULT_PHASE_COLOR


def _phase_from_activity(activity: str) -> tuple[str, str]:
    name = activity.lower()
    if "ev run" in name:
        return "EV Run", "#b79ac7"
    if "frac" in name:
        return "Frac Spot Hire", "#b79ac7"
    if "top hole" in name or "pilot hole" in name:
        phase = PHASE_COLORS["FF00B0F0"]
        return phase["phase"], phase["color"]
    if "demob" in name:
        phase = PHASE_COLORS["FF7030A0"]
        return phase["phase"], phase["color"]
    return "Operational Spot Hire", "#b79ac7"


def _individual_run_phase() -> tuple[str, str]:
    phase = PHASE_COLORS["INDIVIDUAL_RUN"]
    return phase["phase"], phase["color"]


def _completion_phase() -> tuple[str, str]:
    phase = PHASE_COLORS["FF92D050"]
    return phase["phase"], phase["color"]


def _add_missing_baseline_records(records: list[dict]) -> None:
    has_whale_10_completion = any(
        record.get("asset") == "Poseidon (3)"
        and record.get("area") == "Whale"
        and "10" in str(record.get("activity", ""))
        and "completion" in str(record.get("activity", "")).lower()
        for record in records
    )
    if has_whale_10_completion:
        return

    whale_10_deepening = next(
        (
            record for record in records
            if record.get("asset") == "Poseidon (3)"
            and record.get("area") == "Whale"
            and "10" in str(record.get("activity", ""))
            and "deepen" in str(record.get("activity", "")).lower()
        ),
        None,
    )
    if not whale_10_deepening:
        return

    deepening_end = _to_dt(whale_10_deepening.get("end_date"))
    if deepening_end is None:
        return

    phase, color = _completion_phase()
    start = deepening_end + timedelta(days=1)
    end = start + timedelta(days=35)
    records.append({
        "id": _stable_id(SPOT_HIRE_SHEET_NAME, "missing-baseline", "poseidon", "whale-10-completion"),
        "asset": "Poseidon (3)",
        "display_asset": "Poseidon",
        "source_asset_count": 3,
        "vessel_count": 1,
        "area": "Whale",
        "activity": "Whale 10 Completion",
        "phase": phase,
        "color": color,
        "start_date": start.date().isoformat(),
        "end_date": end.date().isoformat(),
        "status": "Planned",
        "notes": "Baseline correction: source workbook includes Whale 10 Deepening but omits Whale 10 Completion.",
        "source": "baseline-correction",
        "record_type": "baseline_correction",
    })


def _phase_from_row_colors(ws, row: int) -> tuple[str, str] | None:
    for col in range(TIMELINE_START_COL, TIMELINE_END_COL + 1):
        phase_color = _fill_rgb(ws.cell(row, col))
        if phase_color in PHASE_COLORS:
            return _phase_from_color(phase_color)
    return None


def _inherited_text(ws, row: int, col: int, last_value: str = "") -> str:
    value = _cell_value(ws, row, col)
    if value is None or str(value).strip() == "":
        return last_value
    return _clean_text(str(value))


def _blue_vessel_anchors(path: Path, timeline_dates: dict[int, datetime]) -> dict[int, list[tuple[int, datetime]]]:
    anchors: dict[int, list[tuple[int, datetime]]] = {}
    namespaces = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    try:
        with zipfile.ZipFile(path) as package:
            if SPOT_HIRE_DRAWING_PATH not in package.namelist():
                return anchors
            root = ET.fromstring(package.read(SPOT_HIRE_DRAWING_PATH))
    except (OSError, ET.ParseError, zipfile.BadZipFile):
        return anchors

    for anchor_name in ("oneCellAnchor", "twoCellAnchor"):
        for anchor in root.findall(f"xdr:{anchor_name}", namespaces):
            blip = anchor.find(".//a:blip", namespaces)
            relation_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed", "") if blip is not None else ""
            if relation_id != BLUE_VESSEL_REL_ID:
                continue
            from_marker = anchor.find("xdr:from", namespaces)
            if from_marker is None:
                continue
            row = int(from_marker.find("xdr:row", namespaces).text) + 1
            col = int(from_marker.find("xdr:col", namespaces).text) + 1
            run_date = timeline_dates.get(col)
            if DATA_START_ROW <= row <= DATA_END_ROW and run_date is not None:
                anchors.setdefault(row, []).append((col, run_date))
    return anchors


def _monthly_forecast_counts(path: Path, timeline_dates: dict[int, datetime]) -> dict[str, int]:
    counts: dict[str, int] = {}
    namespaces = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    try:
        with zipfile.ZipFile(path) as package:
            if SPOT_HIRE_DRAWING_PATH not in package.namelist():
                return counts
            root = ET.fromstring(package.read(SPOT_HIRE_DRAWING_PATH))
    except (OSError, ET.ParseError, zipfile.BadZipFile):
        return counts

    for anchor_name in ("oneCellAnchor", "twoCellAnchor"):
        for anchor in root.findall(f"xdr:{anchor_name}", namespaces):
            from_marker = anchor.find("xdr:from", namespaces)
            if from_marker is None:
                continue
            row = int(from_marker.find("xdr:row", namespaces).text) + 1
            col = int(from_marker.find("xdr:col", namespaces).text) + 1
            if not (DATA_START_ROW <= row <= DATA_END_ROW) or col not in timeline_dates:
                continue
            text = "".join(item.text or "" for item in anchor.findall(".//a:t", namespaces)).strip()
            if not text.isdigit():
                continue
            month_key = timeline_dates[col].strftime("%Y-%m")
            counts[month_key] = max(counts.get(month_key, 0), int(text))
    return counts


def load_spot_hire_forecast_counts(workbook_path: Path | None = None) -> dict[str, int]:
    from openpyxl import load_workbook

    path = workbook_path or WORKBOOK_PATH
    if not path.exists():
        return {}

    workbook = load_workbook(path, data_only=True)
    if SPOT_HIRE_SHEET_NAME not in workbook.sheetnames:
        return {}

    timeline_dates = _timeline_dates(workbook[SPOT_HIRE_SHEET_NAME])
    return _monthly_forecast_counts(path, timeline_dates)


def load_spot_hire_records(workbook_path: Path | None = None) -> tuple[list[dict], str]:
    from openpyxl import load_workbook

    path = workbook_path or WORKBOOK_PATH
    if not path.exists():
        return [], "workbook-not-found"

    workbook = load_workbook(path, data_only=True)
    if SPOT_HIRE_SHEET_NAME not in workbook.sheetnames:
        return [], f"sheet-not-found:{SPOT_HIRE_SHEET_NAME}"

    ws = workbook[SPOT_HIRE_SHEET_NAME]
    timeline_dates = _timeline_dates(ws)
    individual_runs_by_row = _blue_vessel_anchors(path, timeline_dates)
    records: list[dict] = []
    last_asset = ""
    last_area = ""

    for row in range(DATA_START_ROW, min(DATA_END_ROW, ws.max_row) + 1):
        last_asset = _inherited_text(ws, row, 2, last_asset)
        last_area = _inherited_text(ws, row, 3, last_area)
        activity_value = _cell_value(ws, row, 4)
        activity = _clean_text(str(activity_value)) if activity_value is not None else ""
        if not activity:
            continue

        row_start = _to_dt(_cell_value(ws, row, 5))
        row_end_value = _cell_value(ws, row, 6)
        row_end = _to_dt(row_end_value)
        if row_start:
            phase, color = _phase_from_row_colors(ws, row) or _phase_from_activity(activity)
            is_tbd_end = row_end is None and row_end_value is not None and str(row_end_value).strip()
            record_id = _stable_id(SPOT_HIRE_SHEET_NAME, row, "date-columns", activity, phase)
            records.append({
                "id": record_id,
                "asset": last_asset or "Unassigned",
                **_record_identity(last_asset or "Unassigned"),
                "area": last_area,
                "activity": activity,
                "phase": phase,
                "color": color,
                "start_date": row_start.date().isoformat(),
                "end_date": (row_end or row_start).date().isoformat(),
                "status": "Planned",
                "notes": "End Date TBD" if is_tbd_end else "Imported from Start Date / End Date columns",
                "source": "workbook",
                "sheet_row": row,
            })
            individual_phase, individual_color = _individual_run_phase()
            seen_run_cols = set()
            for run_col, run_date in individual_runs_by_row.get(row, []):
                if run_col in seen_run_cols:
                    continue
                seen_run_cols.add(run_col)
                records.append({
                    "id": _stable_id(SPOT_HIRE_SHEET_NAME, row, run_col, "individual-run", activity),
                    "asset": last_asset or "Unassigned",
                    **_record_identity(last_asset or "Unassigned"),
                    "vessel_count": 1,
                    "area": last_area,
                    "activity": f"Additional Run - {activity}",
                    "phase": individual_phase,
                    "color": individual_color,
                    "start_date": run_date.date().isoformat(),
                    "end_date": run_date.date().isoformat(),
                    "status": "Planned",
                    "notes": f"Individual vessel run during {activity}",
                    "source": "workbook",
                    "sheet_row": row,
                    "timeline_start_col": get_column_letter(run_col),
                    "timeline_end_col": get_column_letter(run_col),
                    "parent_activity": activity,
                    "record_type": "individual_run",
                })
            continue

        current_start_col = None
        current_color = None
        row_record_count = 0
        for col in range(TIMELINE_START_COL, TIMELINE_END_COL + 2):
            rgb = _fill_rgb(ws.cell(row, col)) if col <= TIMELINE_END_COL else None
            is_phase_color = rgb in PHASE_COLORS

            if is_phase_color and current_start_col is None:
                current_start_col = col
                current_color = rgb
            elif is_phase_color and rgb == current_color:
                continue
            else:
                if current_start_col is not None and current_color is not None:
                    end_col = col - 1
                    start_date = timeline_dates.get(current_start_col)
                    end_date = _segment_end(end_col, timeline_dates) if end_col in timeline_dates else None
                    if start_date and end_date:
                        phase, color = _phase_from_color(current_color)
                        record_id = _stable_id(SPOT_HIRE_SHEET_NAME, row, current_start_col, end_col, activity, phase)
                        records.append({
                            "id": record_id,
                            "asset": last_asset or "Unassigned",
                            **_record_identity(last_asset or "Unassigned"),
                            "area": last_area,
                            "activity": activity,
                            "phase": phase,
                            "color": color,
                            "start_date": start_date.date().isoformat(),
                            "end_date": end_date.date().isoformat(),
                            "source": "workbook",
                            "sheet_row": row,
                            "timeline_start_col": get_column_letter(current_start_col),
                            "timeline_end_col": get_column_letter(end_col),
                        })
                        row_record_count += 1
                current_start_col = col if is_phase_color else None
                current_color = rgb if is_phase_color else None

        if row_record_count == 0:
            start_date = _to_dt(_cell_value(ws, row, 5))
            end_date = _to_dt(_cell_value(ws, row, 6))
            if start_date and end_date:
                clipped = _clip_to_year(start_date, end_date)
                if clipped:
                    clipped_start, clipped_end = clipped
                    phase, color = _phase_from_activity(activity)
                    record_id = _stable_id(SPOT_HIRE_SHEET_NAME, row, "date-range", activity, phase)
                    records.append({
                        "id": record_id,
                        "asset": last_asset or "Unassigned",
                        **_record_identity(last_asset or "Unassigned"),
                        "area": last_area,
                        "activity": activity,
                        "phase": phase,
                        "color": color,
                        "start_date": clipped_start.date().isoformat(),
                        "end_date": clipped_end.date().isoformat(),
                        "status": "Planned",
                        "notes": "Imported from Start Date / End Date columns",
                        "source": "workbook",
                        "sheet_row": row,
                    })

    _add_missing_baseline_records(records)
    records.sort(key=lambda item: (item["start_date"], item["asset"], item["activity"]))
    return records, f"workbook:{path.name}:{SPOT_HIRE_SHEET_NAME}"
