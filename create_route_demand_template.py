"""
Route Demand Update Workbook Generator
Creates an Excel template pre-filled with current tasks data for easy updates.
"""
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    exit(1)

# Load current data
data_file = Path('data/tasks.json')
with open(data_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "OSV Demand Tracker"

# Define styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Status colors
status_colors = {
    'Complete': '7CB518',
    'In Progress': 'FFD60A',
    'Planned': '00B4D8',
    'On Hold': '6C757D',
    'Cancelled': 'F72585'
}

# Headers - matching expected server.py column mappings
headers = ['Status', 'Asset', 'Project', 'Activity', 'Base Delivery Date', 
           'Offshore Req Date', 'Offloading Complete Date', 
           'Estimated Duration (hrs.)', 'Back to Port', 'Transit Time Back to Port']

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Column widths
column_widths = [12, 15, 25, 50, 16, 16, 18, 18, 16, 20]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Row height for header
ws.row_dimensions[1].height = 30

# Populate with existing tasks
for row_idx, task in enumerate(data.get('tasks', []), 2):
    ws.cell(row=row_idx, column=1, value=task.get('status', 'Planned')).border = thin_border
    ws.cell(row=row_idx, column=2, value=task.get('asset', '')).border = thin_border
    ws.cell(row=row_idx, column=3, value=task.get('project', '')).border = thin_border
    ws.cell(row=row_idx, column=4, value=task.get('activity', '')).border = thin_border
    ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)
    
    # Parse and format dates
    def format_date(date_str):
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str.split()[0], '%Y-%m-%d')
        except:
            return date_str
    
    def format_datetime(date_str):
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        except:
            return date_str
    
    # Start date (Base Delivery Date)
    start_val = format_date(task.get('start_date', ''))
    cell = ws.cell(row=row_idx, column=5, value=start_val)
    cell.border = thin_border
    if isinstance(start_val, datetime):
        cell.number_format = 'YYYY-MM-DD'
    
    # Offshore Req Date
    offshore_start_val = format_datetime(task.get('offshore_start', ''))
    cell = ws.cell(row=row_idx, column=6, value=offshore_start_val)
    cell.border = thin_border
    if isinstance(offshore_start_val, datetime):
        cell.number_format = 'YYYY-MM-DD HH:MM'
    
    # Offloading Complete Date
    offshore_end_val = format_datetime(task.get('offshore_end', ''))
    cell = ws.cell(row=row_idx, column=7, value=offshore_end_val)
    cell.border = thin_border
    if isinstance(offshore_end_val, datetime):
        cell.number_format = 'YYYY-MM-DD HH:MM'
    
    # Duration
    ws.cell(row=row_idx, column=8, value=task.get('duration_hours', 24)).border = thin_border
    
    # Back to Port
    return_val = format_datetime(task.get('return_end', ''))
    cell = ws.cell(row=row_idx, column=9, value=return_val)
    cell.border = thin_border
    if isinstance(return_val, datetime):
        cell.number_format = 'YYYY-MM-DD HH:MM'
    
    # Transit time
    ws.cell(row=row_idx, column=10, value=task.get('transit_hours', 18)).border = thin_border
    
    # Color-code by status
    status = task.get('status', '')
    if status in status_colors:
        fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type='solid')
        ws.cell(row=row_idx, column=1).fill = fill

# Add blank rows for new entries
last_row = len(data.get('tasks', [])) + 2
for row_idx in range(last_row, last_row + 15):
    for col in range(1, 11):
        ws.cell(row=row_idx, column=col).border = thin_border
    # Default values for new rows
    ws.cell(row=row_idx, column=1, value='Planned')
    ws.cell(row=row_idx, column=8, value=24)  # Default duration
    ws.cell(row=row_idx, column=10, value=18)  # Default transit

# Add a reference sheet
ws2 = wb.create_sheet("Reference")
ws2['A1'] = "Valid Statuses"
ws2['A1'].font = Font(bold=True)
for i, status in enumerate(status_colors.keys(), 2):
    ws2.cell(row=i, column=1, value=status)
    ws2.cell(row=i, column=1).fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type='solid')

ws2['C1'] = "Column Descriptions"
ws2['C1'].font = Font(bold=True)
descriptions = [
    ('Status', 'Complete, In Progress, Planned, On Hold, Cancelled'),
    ('Asset', 'Platform/rig name (Pontus, Poseidon, etc.)'),
    ('Project', 'Well/campaign name'),
    ('Activity', 'Description of cargo/scope (e.g., GL 01 OSV / 24 hrs / ...)'),
    ('Base Delivery Date', 'Sail date from port (YYYY-MM-DD)'),
    ('Offshore Req Date', 'Arrival at rig (YYYY-MM-DD HH:MM)'),
    ('Offloading Complete Date', 'Departure from rig (YYYY-MM-DD HH:MM)'),
    ('Estimated Duration', 'Hours at offshore location'),
    ('Back to Port', 'Arrival back at port (YYYY-MM-DD HH:MM)'),
    ('Transit Time', 'Transit hours from offshore to port')
]
for i, (col_name, desc) in enumerate(descriptions, 2):
    ws2.cell(row=i, column=3, value=col_name).font = Font(bold=True)
    ws2.cell(row=i, column=4, value=desc)

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['C'].width = 25
ws2.column_dimensions['D'].width = 50

# Save
output_file = f'Route_Demand_Update_{datetime.now().strftime("%Y%m%d")}.xlsx'
wb.save(output_file)
print(f"Created: {output_file}")
print(f"Tasks: {len(data.get('tasks', []))}")
print("\nInstructions:")
print("1. Edit the 'OSV Demand Tracker' sheet")
print("2. Add new rows or modify existing ones")
print("3. Make sure 'Activity' column has descriptions")
print("4. Save the file")
print("5. Upload via the OSV Demand Scheduler web page")
